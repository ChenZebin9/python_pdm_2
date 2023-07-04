import xlrd
from openpyxl import load_workbook
from abc import abstractmethod, ABCMeta


class ExcelHandler( metaclass=ABCMeta ):

    @abstractmethod
    def get_sheets_name(self):
        pass

    @abstractmethod
    def get_max_rows(self):
        pass

    @abstractmethod
    def get_datas(self, sheet_name):
        pass


class ExcelHandler3( ExcelHandler ):
    """解决操作xlsx(Excel 2010)文件的问题"""

    def __init__(self, file_name):
        self.__book = load_workbook( file_name, data_only=True )
        pass

    def close(self):
        self.__book.close()

    def get_sheets_name(self):
        t = self.__book.sheetnames
        return t

    def get_max_rows(self, sheet_name):
        sh = self.__book[sheet_name]
        return sh.max_row

    def get_datas(self, sheet_name):
        sh = self.__book[sheet_name]
        columns = []
        data_s = {}
        for col in sh.columns:
            first_cell = col[0]
            if sh.column_dimensions[first_cell.column_letter].hidden:
                continue
            col_name = first_cell.value
            columns.append( col_name )
            column_data = []
            for c in col[1:]:
                if sh.row_dimensions[c.row].hidden:
                    continue
                vv = c.value
                if vv is None:
                    vv = ''
                column_data.append( (c.row, vv) )
            data_s[col_name] = column_data
        return columns, data_s


class ExcelHandler2( ExcelHandler ):
    """ 使用 xlrd 速度比较快 """

    def __init__(self, file_name):
        # 要加 formatting_info=True 才能获得单元格的格式
        self.__book = xlrd.open_workbook( file_name, formatting_info=True )

    def close(self):
        self.__book.release_resources()

    def get_sheets_name(self):
        return self.__book.sheet_names()

    def get_max_rows(self, sheet_name):
        ss = self.__book.sheet_by_name( sheet_name )
        return ss.nrows + 1

    def get_datas(self, sheet_name):
        from xlrd import xldate_as_datetime
        ss = self.__book.sheet_by_name( sheet_name )
        columns = []
        datas = {}
        cinfomap = ss.colinfo_map
        rinfomap = ss.rowinfo_map
        for col_index in range( ss.ncols ):
            if cinfomap.get( col_index, 0 ) and cinfomap[col_index].hidden == 1:
                continue
            tcell = ss.cell( 0, col_index )
            col_name = tcell.value
            columns.append( col_name )
            data = []
            for row_index in range( 1, ss.nrows ):
                if rinfomap.get( row_index, 0 ) and rinfomap[row_index].hidden == 1:
                    continue
                tcell = ss.cell( row_index, col_index )
                if tcell.ctype == 3:  # 处理日期数据
                    date_value = xldate_as_datetime(tcell.value, 0)
                    vv = date_value.strftime( '%Y-%m-%d' )
                else:
                    vv = tcell.value
                if vv is None:
                    vv = ''
                # 增加行号的数据
                data.append( (row_index + 1, str( vv )) )
            datas[col_name] = data
        return columns, datas
