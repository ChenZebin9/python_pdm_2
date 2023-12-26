# coding=gbk
import datetime
import os.path

import xlwings as xw
from PIL import Image
from PyQt5.QtCore import Qt, QDate, QPoint, QThread, pyqtSignal
from PyQt5.QtGui import QCursor, QPixmap, QColor, QPalette, QBrush
from PyQt5.QtWidgets import QMainWindow, QTreeWidget, QFrame, QHBoxLayout, QLabel, QTableWidget, \
    QVBoxLayout, QSplitter, QTreeWidgetItem, QTableWidgetItem, QSpinBox, QInputDialog, QAbstractItemView, QMenu, \
    QDoubleSpinBox, QFileDialog, QMessageBox, QDialog, QHeaderView, QWidget, QLineEdit, QListWidget, QListWidgetItem, \
    QTabWidget, QPushButton
from openpyxl import load_workbook

from Part import Part
from db.DatabaseHandler import DatabaseHandler
from excel.ExcelHandler import ExcelHandler3, ExcelHandler2
from ui.BlankDialog import Ui_Dialog as BlankDialog
from ui.NImportSettingDialog import NImportSettingDialog
from ui3.NCreatePickBillDialog import NCreatePickBillDialog
from ui4.AssemblyToolWindow import Ui_MainWindow
from Com import is_used


class NoWheelPinBox(QDoubleSpinBox):
    """
    �������ι��������������
    """

    def __init__(self, parent=None):
        super(NoWheelPinBox, self).__init__(parent)

    def wheelEvent(self, e) -> None:
        # ���ι����¼�
        pass


# noinspection PyUnresolvedReferences
class QtyComWidget(QWidget):

    def __init__(self, qty, value, parent=None, min_value=0., max_value=99., a_qty=0.):
        """
        ��ʾ�ѷ�������ϵ���Ϣ
        :param qty: ���п����
        :param value: ���䵽����Ŀ������
        :param parent:
        :param min_value: ��С����
        :param max_value: �������
        :param a_qty: �ѷ��䵽������Ŀ������
        """
        super(QtyComWidget, self).__init__(parent)
        self.__qty = qty
        v_layout = QVBoxLayout(self)
        v_layout.setSpacing(0)
        v_layout.setContentsMargins(0, 0, 0, 0)
        the_line_edit = QLabel('{0:.2f}'.format(qty))
        the_line_edit.setAlignment(Qt.AlignCenter)
        the_line_edit_2 = QLabel('{0:.2f}'.format(a_qty))
        the_line_edit_2.setAlignment(Qt.AlignCenter)
        self.menu_4_line_2 = QMenu(the_line_edit_2)
        self.see_detail_action = self.menu_4_line_2.addAction('����')
        self.see_detail_action.triggered.connect(self.see_detail)
        the_line_edit_2.setContextMenuPolicy(Qt.CustomContextMenu)
        the_line_edit_2.customContextMenuRequested.connect(self.on_right_click)
        v_layout.addWidget(the_line_edit)
        v_layout.addWidget(the_line_edit_2)
        self.__detail_assigned_message = None
        self.__the_spin_box = NoWheelPinBox(self)
        self.__the_spin_box.setMinimum(min_value)
        self.__the_spin_box.setMaximum(max_value if max_value > 0.0 else 0.0)
        self.__the_spin_box.setValue(value if value >= 0.0 else 0.0)
        self.__the_spin_box.setAlignment(Qt.AlignCenter)
        v_layout.addWidget(self.__the_spin_box)
        self.setLayout(v_layout)
        # self.setFixedSize(64, 64)

    def set_assigned_detail_data(self, data_list, _index):
        """
        ������ϸ������Ԥ������Ϣ
        :param data_list:
        :param _index: 1 - �е£�2 - ���֣�3 - �ֳ�
        :return:
        """
        tt = []
        for d in data_list:
            qty = d[_index + 1]
            if qty <= 0.0:
                continue
            tt.append(f'{d[0]} -> {d[1]} -> {qty:.2f}')
        self.__detail_assigned_message = '\n'.join(tt) if len(tt) > 0 else None

    def see_detail(self):
        QMessageBox.information(None, '�ѷ�������', self.__detail_assigned_message)

    def on_right_click(self):
        if self.__detail_assigned_message is None:
            return
        self.menu_4_line_2.exec_(QCursor.pos())

    def value(self):
        return self.__the_spin_box.value()

    def set_value(self, v):
        self.__the_spin_box.setValue(v)

    def get_qty(self):
        return self.__qty


class ProgressStructPanel(QFrame):

    def __init__(self, parent, database: DatabaseHandler):
        super(ProgressStructPanel, self).__init__(parent)
        self.__parent = parent
        self.__database = database
        self.__selected_progress = []
        # ��ʾ���̽ṹ�������б�
        self.progressStructTreeWidget = QTreeWidget(self)
        self.setup_ui()
        self.data_init()

    def setup_ui(self):
        v_box = QVBoxLayout(self)
        v_box.addWidget(QLabel('����ṹ'))
        v_box.addWidget(self.progressStructTreeWidget)
        self.setLayout(v_box)
        self.progressStructTreeWidget.setSelectionMode(QAbstractItemView.ExtendedSelection)
        # ��Ӧ����
        self.progressStructTreeWidget.itemExpanded.connect(self.__when_item_expand)

    def data_init(self):
        self.progressStructTreeWidget.setColumnCount(1)
        self.progressStructTreeWidget.header().hide()
        struct_bom_s = Part.get_parts(self.__database, part_id=5090)
        for struct_bom in struct_bom_s:
            children_part = struct_bom.get_children(self.__database)
            root_items = []
            for c in children_part:
                p: Part = c[1]
                t = p.get_specified_tag(self.__database, '���')
                if t == '�ĵ�' or t == 'ͼֽ':
                    continue
                r_node = QTreeWidgetItem(self.progressStructTreeWidget)
                r_node.setText(0, '{0:04d} {1}'.format(p.get_part_id(), p.name))
                r_node.setData(0, Qt.UserRole, (p, True, 1.))
                p_c = p.get_children(self.__database)
                for pc in p_c:
                    c_node = QTreeWidgetItem()
                    p_cc = pc[1]
                    t1 = p_cc.get_specified_tag(self.__database, '���')
                    if t1 == '�ĵ�' or t1 == 'ͼֽ':
                        continue
                    # װ���嵥ͳ��ʱ���˲���<ʵ������>
                    c_node.setText(0, '{0:04d} {1} x {2}'.format(p_cc.get_part_id(), p_cc.name, round(c[3])))
                    t = p_cc.get_specified_tag(self.__database, '��Դ')
                    c_flag = False
                    if t is not None and t == 'װ��':
                        c_flag = True
                    c_node.setData(0, Qt.UserRole, (p_cc, c_flag, c[3]))
                    r_node.addChild(c_node)
                root_items.append(r_node)
            self.progressStructTreeWidget.addTopLevelItems(root_items)
        self.progressStructTreeWidget.resizeColumnToContents(0)

    def get_selected_item(self):
        self.__selected_progress.clear()
        selected_items = self.progressStructTreeWidget.selectedItems()
        if len(selected_items) < 1:
            return None
        ss = selected_items
        for s in ss:
            s.setExpanded(True)
            self.__get_progress(s, 1.0)
        return self.__selected_progress

    def __get_progress(self, item: QTreeWidgetItem, qty):
        dd = item.data(0, Qt.UserRole)
        item_qty = dd[2]
        if item.childCount() > 0:
            for i in range(item.childCount()):
                ci = item.child(i)
                ci.setExpanded(True)
                self.__get_progress(ci, qty * item_qty)
        else:
            p = item.data(0, Qt.UserRole)[0]
            self.__selected_progress.append((p, qty * item_qty))

    def __when_item_expand(self, item: QTreeWidgetItem):
        c_count = item.childCount()
        for i in range(0, c_count):
            cc = item.child(i)
            dd = cc.data(0, Qt.UserRole)
            if dd[1]:
                continue
            pp = dd[0]
            cc.setData(0, Qt.UserRole, (pp, True, dd[2]))
            children = pp.get_children(self.__database)
            if children is None:
                continue
            nodes = []
            for c in children:
                node = QTreeWidgetItem()
                p = c[1]
                tt = p.get_specified_tag(self.__database, '���')
                if tt == '�ĵ�' or tt == 'ͼֽ':
                    continue
                node.setText(0, '{0:04d} {1} x {2}'.format(p.get_part_id(), p.name, round(c[2])))
                t = p.get_specified_tag(self.__database, '��Դ')
                c_flag = False
                if t is not None and (t == 'װ��' or t == '����'):
                    c_flag = True
                else:
                    t1 = p.get_specified_tag(self.__database, '���ϲ���')
                    if t1 is not None and (t1 == 'PST'):
                        c_flag = True
                node.setData(0, Qt.UserRole, (p, c_flag, c[2]))
                nodes.append(node)
            cc.addChildren(nodes)
        self.progressStructTreeWidget.resizeColumnToContents(0)


# noinspection PyUnresolvedReferences
class SelectedProcessPanel(QFrame):

    def __init__(self, parent=None):
        super(SelectedProcessPanel, self).__init__(parent)
        self.__parent = parent
        # ��ѡ�����̱��
        self.selectedProcessTableWidget = QTableWidget(self)
        # �Ҽ��˵�
        self.__menu_4_table = QMenu(parent=self.selectedProcessTableWidget)
        self.__delete_item_action = self.__menu_4_table.addAction('ɾ��')
        self.__clean_item_action = self.__menu_4_table.addAction('���')
        self.setup_ui()

    def setup_ui(self):
        v_box = QVBoxLayout(self)
        v_box.addWidget(QLabel('��ѡ����'))
        v_box.addWidget(self.selectedProcessTableWidget)
        self.setLayout(v_box)
        self.selectedProcessTableWidget.setColumnCount(3)
        self.selectedProcessTableWidget.setHorizontalHeaderLabels(('���', '����', '����'))
        self.selectedProcessTableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.selectedProcessTableWidget.setContextMenuPolicy(Qt.CustomContextMenu)
        self.selectedProcessTableWidget.customContextMenuRequested.connect(self.__on_custom_context_menu_requested)
        self.__delete_item_action.triggered.connect(self.__delete_selected)
        self.__clean_item_action.triggered.connect(self.__clean_selected)

    def __clean_selected(self):
        r_c = self.selectedProcessTableWidget.rowCount()
        for i in range(r_c):
            self.selectedProcessTableWidget.removeRow(0)

    def __delete_selected(self):
        items = self.selectedProcessTableWidget.selectedItems()
        r_list = []
        for i in items:
            r_i = i.row()
            if r_i not in r_list:
                r_list.append(r_i)
        r_list.sort(reverse=True)
        for r in r_list:
            self.selectedProcessTableWidget.removeRow(r)

    def __on_custom_context_menu_requested(self, pos):
        if self.selectedProcessTableWidget.rowCount() < 1:
            return
        item = self.selectedProcessTableWidget.itemAt(pos)
        self.__delete_item_action.setVisible(item is not None)
        self.__menu_4_table.exec_(QCursor.pos())

    def get_all_item(self):
        r = []
        c = self.selectedProcessTableWidget.rowCount()
        for i in range(c):
            item_p = self.selectedProcessTableWidget.item(i, 0)
            p = item_p.data(Qt.UserRole)
            item_w = self.selectedProcessTableWidget.cellWidget(i, 2)
            q = item_w.value()
            r.append((p, q))
        return r

    def add_row(self, data):
        r_c = self.selectedProcessTableWidget.rowCount()
        neu_r = len(data)
        self.selectedProcessTableWidget.setRowCount(r_c + neu_r)
        i = r_c
        for d in data:
            p: Part = d[0]
            qty = d[1]
            part_id_item = QTableWidgetItem()
            part_id_item.setData(Qt.DisplayRole, '{0:04d}'.format(p.get_part_id()))
            part_id_item.setFlags(part_id_item.flags() & (~Qt.ItemIsEditable))
            part_id_item.setTextAlignment(Qt.AlignHCenter)
            part_id_item.setData(Qt.UserRole, p)
            self.selectedProcessTableWidget.setItem(i, 0, part_id_item)
            description_item = QTableWidgetItem()
            des = [p.name]
            if p.description is not None:
                des.append(p.description)
            description_item.setData(Qt.DisplayRole, ' '.join(des))
            description_item.setFlags(description_item.flags() & (~Qt.ItemIsEditable))
            description_item.setTextAlignment(Qt.AlignHCenter)
            self.selectedProcessTableWidget.setItem(i, 1, description_item)
            qty_item = QSpinBox()
            qty_item.setValue(qty)
            qty_item.setAlignment(Qt.AlignHCenter)
            self.selectedProcessTableWidget.setCellWidget(i, 2, qty_item)
            i += 1
        QTableWidget.resizeColumnsToContents(self.selectedProcessTableWidget)
        QTableWidget.resizeRowsToContents(self.selectedProcessTableWidget)
        self.selectedProcessTableWidget.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)


# noinspection PyUnresolvedReferences
class MaterialTablePanel(QFrame):
    """
    2023-12-26 ��Ϊ�������ֳ�����ŷ���е²ִ�Ϊ���ϲο�
    """
    # __header = ('ͼʾ', '�����', '����', '����', '��λ',
    #             '�е�ERP���ϱ��', '���ÿ��', '��������', '����ERP���ϱ��', '���ÿ��', '��������',
    #             '�����ֳ����', '��������', '���ϲ���', '�ɹ��ƻ�')
    __header = ('ͼʾ', '�����', '����', '����', '��λ',
                '��ŷERP���ϱ��', '���ÿ��', '��������', '�е�ERP���ϱ��', '���ÿ��', '��������',
                '�����ֳ����', '��������', '���ϲ���', '�ɹ��ƻ�')

    def __init__(self, parent, database, mode=0, assigned_data=None):
        """
        ������ʾ���
        :param parent:
        :param database:
        :param mode: 0=һ��ģʽ��1=�Ի���ģʽ
        """
        super(MaterialTablePanel, self).__init__(parent)
        self.__parent = parent
        self.__database: DatabaseHandler = database
        self.__mode = mode
        self.__local_assigned_data = assigned_data
        self.__spin_box_dict = {}  # ��Ӧ��Ҫ��Ԫ��1�����ֵ�
        # ����嵥���
        self.materialTableWidget = QTableWidget(self)
        self.__sort_flags = {}
        # ���ϱ����Ҽ��˵�
        self.__menu_4_table = QMenu(parent=self.materialTableWidget)
        self.__replace_item_action = self.__menu_4_table.addAction('�滻')
        self.__delete_item_action = self.__menu_4_table.addAction('ɾ��')
        self.__insert_item_action = self.__menu_4_table.addAction('����')
        self.__append_item_action = self.__menu_4_table.addAction('����')
        self.__refresh_item_action = self.__menu_4_table.addAction('ˢ�¿��')
        # �Ҽ��˵�ʱ��������������
        self.__click_row = -1
        # ���ͼ�δ�������״̬
        self.in_pick_up_mode = False
        self.pick_up_list_file = ''
        # �嵥û����ı��
        self.table_need_save = False
        self.setup_ui()

    def setup_ui(self):
        v_box = QVBoxLayout(self)
        v_box.addWidget(self.materialTableWidget)
        self.setLayout(v_box)
        self.set_header()
        self.materialTableWidget.setSelectionMode(QAbstractItemView.SingleSelection)
        self.materialTableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        header_goods = self.materialTableWidget.horizontalHeader()
        header_goods.sectionClicked.connect(self.__sort_by_column)
        if self.__mode == 0:
            self.materialTableWidget.setContextMenuPolicy(Qt.CustomContextMenu)
            self.materialTableWidget.customContextMenuRequested.connect(self.__on_custom_context_menu_requested)
            self.__replace_item_action.triggered.connect(self.__replace_item)
            self.__delete_item_action.triggered.connect(self.__delete_item)
            self.__insert_item_action.triggered.connect(lambda: self.__insert_item(self.__click_row))
            self.__append_item_action.triggered.connect(self.__append_item)
            self.__refresh_item_action.triggered.connect(self.__refresh_item)

    def __on_custom_context_menu_requested(self, pos: QPoint):
        if self.sender() is self.materialTableWidget:
            column = self.materialTableWidget.columnAt(pos.x())
            self.__click_row = self.materialTableWidget.rowAt(pos.y())
            on_item = column >= 0 and self.__click_row >= 0
            self.__replace_item_action.setVisible(on_item and not self.in_pick_up_mode)
            self.__delete_item_action.setVisible(on_item and not self.in_pick_up_mode)
            self.__insert_item_action.setVisible(on_item and not self.in_pick_up_mode)
            self.__append_item_action.setVisible((not on_item) and not self.in_pick_up_mode)
            self.__menu_4_table.exec_(QCursor.pos())

    # noinspection PyTypeChecker
    def __change_qty_2(self):
        """
        ���� __change_qty������ŷΪ�������ϲο���2023-12-26
        :return:
        """
        w = self.sender()
        id_cell: QTableWidgetItem = self.__spin_box_dict[w]
        neu_value = w.value()
        u_data = id_cell.data(Qt.UserRole)
        neu_data = (u_data[0], u_data[1], u_data[2], neu_value)
        if neu_value > u_data[1] + u_data[2]:
            id_cell.setBackground(QColor(255, 0, 0))
        else:
            id_cell.setBackground(QColor(255, 255, 255))
        id_cell.setData(Qt.UserRole, neu_data)
        # �������Ϸ���
        remain_qty = neu_value
        r = self.materialTableWidget.indexFromItem(id_cell).row()
        site_storing_w = self.materialTableWidget.cellWidget(r, 11)
        if site_storing_w is not None:
            w: QtyComWidget = site_storing_w
            q = w.get_qty()
            if q >= remain_qty:
                vv = remain_qty
            else:
                vv = q
            remain_qty -= vv
            w.set_value(vv)
        jo_storing_w = self.materialTableWidget.cellWidget(r, 6)
        if jo_storing_w is not None:
            w: QtyComWidget = jo_storing_w
            q = w.get_qty()
            if q >= remain_qty:
                vv = remain_qty
            else:
                vv = q
            remain_qty -= vv
            w.set_value(vv)
        zd_storing_w = self.materialTableWidget.cellWidget(r, 9)
        if jl_storing_w is not None:
            w: QtyComWidget = zd_storing_w
            q = w.get_qty()
            if q >= remain_qty:
                vv = remain_qty
            else:
                vv = q
            remain_qty -= vv
            w.set_value(vv)
        if self.in_pick_up_mode:
            self.table_need_save = True
            self.__parent.set_save_btn(True)

    def __change_qty(self):
        w = self.sender()
        id_cell: QTableWidgetItem = self.__spin_box_dict[w]
        neu_value = w.value()
        u_data = id_cell.data(Qt.UserRole)
        neu_data = (u_data[0], u_data[1], u_data[2], neu_value)
        if neu_value > u_data[1] + u_data[2]:
            id_cell.setBackground(QColor(255, 0, 0))
        else:
            id_cell.setBackground(QColor(255, 255, 255))
        id_cell.setData(Qt.UserRole, neu_data)
        # �������Ϸ���
        remain_qty = neu_value
        r = self.materialTableWidget.indexFromItem(id_cell).row()
        site_storing_w = self.materialTableWidget.cellWidget(r, 11)
        if site_storing_w is not None:
            w: QtyComWidget = site_storing_w
            q = w.get_qty()
            if q >= remain_qty:
                vv = remain_qty
            else:
                vv = q
            remain_qty -= vv
            w.set_value(vv)
        zd_storing_w = self.materialTableWidget.cellWidget(r, 6)
        if zd_storing_w is not None:
            w: QtyComWidget = zd_storing_w
            q = w.get_qty()
            if q >= remain_qty:
                vv = remain_qty
            else:
                vv = q
            remain_qty -= vv
            w.set_value(vv)
        jl_storing_w = self.materialTableWidget.cellWidget(r, 9)
        if jl_storing_w is not None:
            w: QtyComWidget = jl_storing_w
            q = w.get_qty()
            if q >= remain_qty:
                vv = remain_qty
            else:
                vv = q
            remain_qty -= vv
            w.set_value(vv)
        if self.in_pick_up_mode:
            self.table_need_save = True
            self.__parent.set_save_btn(True)

    def __insert_item(self, row):
        if len(self.__parent.assigned_material_dict) < 1:
            self.__parent.get_assigned_materials()
        # ���루��ǰ���ӣ�һ��
        txt, ok = QInputDialog.getText(self, '������Ŀ', '�����', QLineEdit.Normal, '�����Ŀ���Կո������')
        if not ok:
            return
        tt = txt.strip().split(' ')
        try:
            r = row
            for t in tt:
                part_number = int(t)
                p = Part.get_parts(self.__database, part_id=part_number)[0]
                self.materialTableWidget.insertRow(r)
                self.__fill_one_row(p, 0.0, r)
                r += 1
            QTableWidget.resizeColumnsToContents(self.materialTableWidget)
        except Exception as ex:
            QMessageBox.warning(self, '����', ex.__str__())

    def fill_pick_datas(self, pick_datas):
        # ���������ݽ������
        r = 0
        for p_d in pick_datas:
            the_part = Part.get_parts(self.__database, part_id=p_d[0])[0]
            qty = p_d[1]
            pick_data = (p_d[2], p_d[3], p_d[4])
            self.materialTableWidget.insertRow(r)
            self.__fill_one_row(the_part, qty, r, pick_data)
            r += 1
        QTableWidget.resizeColumnsToContents(self.materialTableWidget)

    def __append_item(self):
        # ���ӣ���������ӣ�һ��
        row_max = self.materialTableWidget.rowCount()
        self.__insert_item(row_max)

    def __refresh_item(self):
        # ˢ��������Ϣ
        parent_w: MaterialTab = self.__parent
        parent_w.get_assigned_materials()
        row_count = self.materialTableWidget.rowCount()
        for i in range(row_count):
            p_item = self.materialTableWidget.item(i, 1)
            r_data = p_item.data(Qt.UserRole)
            p: Part = r_data[0]
            # �ѷ������ϵ���ϸ��Ϣ
            a_dict = self.__parent.detail_assigned_dict
            part_a_detail_list = a_dict[p.get_part_id()] if p.get_part_id() in a_dict else None
            # �ִ���Ϣ
            qty_in_storing = 0.  # �ֿ��������
            qty_in_site = 0.  # �ֳ�������
            storing_data_s = self.__database.get_storing(part_id=p.get_part_id())
            zd_erp_id = p.get_specified_tag(self.__database, '�����е�ERP���ϱ���')
            jl_erp_id = p.get_specified_tag(self.__database, '��������ERP���ϱ���')
            w: NoWheelPinBox = self.materialTableWidget.cellWidget(i, 3)
            qty = w.value()
            remain_qty = qty  # ʣ�������
            # �ѷ��������
            s1 = 0.0
            s2 = 0.0
            s3 = 0.0
            if p.get_part_id() in parent_w.assigned_material_dict:
                ss = parent_w.assigned_material_dict[p.get_part_id()]
                s1 = ss[0]  # �е²ֿ�
                s2 = ss[1]  # ���ֲֿ�
                s3 = ss[2]  # �ֳ�
            # �ֳ��ִ���Ϣ
            if storing_data_s is not None:
                s_qty = 0.
                u_date = None
                for r_s in storing_data_s:
                    if r_s[1] != 'A':
                        continue
                    s_qty += r_s[2]
                    if (u_date is not None and r_s[3] > u_date) or u_date is None:
                        u_date = r_s[3]
                s_o_qty = s_qty
                s_qty -= s3
                qty_in_site += s_qty
                if u_date is not None and s_o_qty > 0.:
                    if s_qty >= remain_qty:
                        vv = remain_qty
                    else:
                        vv = s_qty
                    remain_qty -= vv
                    qty_w = QtyComWidget(s_qty + s3, vv, parent=self.materialTableWidget, max_value=s_qty, a_qty=s3)
                    if part_a_detail_list is not None:
                        qty_w.set_assigned_detail_data(part_a_detail_list, 3)
                    self.materialTableWidget.setCellWidget(i, 11, qty_w)
                    u_date_cell = QTableWidgetItem()
                    if type(u_date) == str:
                        u_date = datetime.datetime.fromisoformat(u_date)
                    u_date_cell.setData(Qt.DisplayRole, u_date.strftime('%Y/%m/%d'))
                    self.materialTableWidget.setItem(i, 12, u_date_cell)
            # �е�ERP�ִ���Ϣ
            if zd_erp_id != '':
                erp_id_cell = QTableWidgetItem()
                erp_id_cell.setData(Qt.DisplayRole, zd_erp_id)
                self.materialTableWidget.setItem(i, 5, erp_id_cell)
                s_qty = 0.
                u_date = None
                if storing_data_s is not None:
                    for r_s in storing_data_s:
                        if r_s[1] != 'D' and r_s[1] != 'E':
                            continue
                        s_qty += r_s[2]
                        if (u_date is not None and r_s[3] > u_date) or u_date is None:
                            u_date = r_s[3]
                    s_o_qty = s_qty
                    s_qty -= s1
                    qty_in_storing += s_qty
                    if u_date is not None and s_o_qty > 0.:
                        if s_qty >= remain_qty:
                            vv = remain_qty
                        else:
                            vv = s_qty
                        remain_qty -= vv
                        qty_w = QtyComWidget(s_qty + s1, vv, parent=self.materialTableWidget, max_value=s_qty, a_qty=s1)
                        if part_a_detail_list is not None:
                            qty_w.set_assigned_detail_data(part_a_detail_list, 1)
                        self.materialTableWidget.setCellWidget(i, 6, qty_w)
                        u_date_cell = QTableWidgetItem()
                        if type(u_date) == str:
                            u_date = datetime.datetime.fromisoformat(u_date)
                        u_date_cell.setData(Qt.DisplayRole, u_date.strftime('%Y/%m/%d'))
                        self.materialTableWidget.setItem(i, 7, u_date_cell)
            # ����ERP�ִ���Ϣ
            if jl_erp_id != '':
                erp_id_cell = QTableWidgetItem()
                erp_id_cell.setData(Qt.DisplayRole, jl_erp_id)
                self.materialTableWidget.setItem(i, 8, erp_id_cell)
                s_qty = 0.
                u_date = None
                if storing_data_s is not None:
                    for r_s in storing_data_s:
                        if r_s[1] != 'F':
                            continue
                        s_qty += r_s[2]
                        if (u_date is not None and r_s[3] > u_date) or u_date is None:
                            u_date = r_s[3]
                    s_o_qty = s_qty
                    s_qty -= s2
                    qty_in_storing += s_qty
                    if u_date is not None and s_o_qty > 0.:
                        if s_qty >= remain_qty:
                            vv = remain_qty
                        else:
                            vv = s_qty
                        remain_qty -= vv
                        qty_w = QtyComWidget(s_qty + s2, vv, parent=self.materialTableWidget, max_value=s_qty, a_qty=s2)
                        if part_a_detail_list is not None:
                            qty_w.set_assigned_detail_data(part_a_detail_list, 1)
                        self.materialTableWidget.setCellWidget(i, 9, qty_w)
                        u_date_cell = QTableWidgetItem()
                        if type(u_date) == str:
                            u_date = datetime.datetime.fromisoformat(u_date)
                        u_date_cell.setData(Qt.DisplayRole, u_date.strftime('%Y/%m/%d'))
                        self.materialTableWidget.setItem(i, 10, u_date_cell)
            it = self.materialTableWidget.item(i, 1)  # ���ڴ洢���ݵĵ�Ԫ
            # ���������Աȣ��ı��е���ɫ
            if qty > qty_in_storing + qty_in_site:
                it.setBackground(QColor(255, 0, 0))
            else:
                it.setBackground(QColor(255, 255, 255))
            it.setData(Qt.UserRole, (p, qty_in_storing, qty_in_site, qty))
            self.materialTableWidget.setRowHeight(i, 66)
        QMessageBox.information(self, '', '������ϣ�')
        self.table_need_save = True
        self.__parent.set_save_btn(True)

    def __delete_item(self):
        item_r = self.materialTableWidget.currentRow()
        p_item = self.materialTableWidget.item(item_r, 1)
        r_data = p_item.data(Qt.UserRole)
        the_part: Part = r_data[0]
        resp = QMessageBox.question(self.__parent, 'ȷ��', 'ȷ��Ҫ����ɾ����', QMessageBox.Yes | QMessageBox.No,
                                    QMessageBox.No)
        if resp == QMessageBox.Yes:
            self.materialTableWidget.removeRow(item_r)
            self.__parent.show_status_message(f'�ո�ɾ�������{the_part.part_id}��')

    def __replace_item(self):
        item_r = self.materialTableWidget.currentRow()
        p_item = self.materialTableWidget.item(item_r, 1)
        r_data = p_item.data(Qt.UserRole)
        the_part: Part = r_data[0]
        w: NoWheelPinBox = self.materialTableWidget.cellWidget(item_r, 3)
        qty = w.value()
        identical_parts = self.__database.get_identical_parts(the_part.get_part_id())
        if identical_parts is None or len(identical_parts[0]) < 1:
            self.__parent.show_status_message('û�пɴ�������ϣ�', 3)
            return
        part_dict = {}
        for p_id in identical_parts[0]:
            ps = Part.get_parts(self.__database, part_id=p_id)
            if len(ps) > 0:
                p: Part = ps[0]
                part_dict[p] = qty
        if len(identical_parts[0]) > 0:
            dialog = PartSelectDialog(self.__parent, self.__database, part_dict,
                                      assigned_data=self.__parent.assigned_material_dict,
                                      purchase_plan=self.__parent.purchase_plan)
            dialog.setWindowTitle(f'ѡ�����<{the_part.name}>�����')
            resp = dialog.exec_()
            if resp != QDialog.Accepted:
                return
            r = dialog.selected_data
            if r[1] >= qty:
                self.__fill_one_row(r[0], qty, item_r)
                if w in self.__spin_box_dict:
                    self.__spin_box_dict.pop(w)
                neu_p: Part = r[0]
                self.__parent.show_status_message(f'{neu_p.part_id}������{the_part.part_id}��')
            else:
                remain_qty = qty - r[1]
                w.setValue(remain_qty)
                self.materialTableWidget.insertRow(item_r + 1)
                self.__fill_one_row(r[0], r[1], item_r + 1)
            QTableWidget.resizeColumnsToContents(self.materialTableWidget)

    def __sort_by_column(self, column_index):
        # ���������Ӧ����
        sort_flags = False
        if column_index == 0 or column_index == 3:
            return
        if column_index in self.__sort_flags:
            sort_flags = self.__sort_flags[column_index]
        self.materialTableWidget.sortByColumn(column_index, Qt.AscendingOrder if sort_flags else Qt.DescendingOrder)
        self.__sort_flags[column_index] = ~sort_flags

    def set_header(self):
        self.materialTableWidget.setColumnCount(len(MaterialTablePanel.__header))
        self.materialTableWidget.setHorizontalHeaderLabels(MaterialTablePanel.__header)

    def __fill_one_row_2(self, p: Part, qty, row_index, pick_datas=None):
        """
        ����__fill_one_row������һ�У�2023-12-26
        :param p:
        :param qty:
        :param row_index:
        :param pick_datas: (��ŷ�ѷ��䣬 �е��ѷ��䣬 �ֳ��ѷ���)
        :return:
        """
        i = row_index
        original_item = self.materialTableWidget.item(row_index, 2)
        if original_item is not None:
            # ɾ��ԭ����Ϣ
            self.materialTableWidget.insertRow(i)
            self.materialTableWidget.removeRow(i + 1)
        # ͼʾ
        img_data = self.__database.get_thumbnail_2_part(p.get_part_id())
        if img_data is not None:
            img = QPixmap()
            img.loadFromData(img_data)
            n_img = img.scaled(64, 64, aspectRatioMode=Qt.KeepAspectRatio)
            img_label_w = QLabel()
            img_label_w.setPixmap(n_img)
            img_label_w.setAlignment(Qt.AlignCenter)
            self.materialTableWidget.setCellWidget(i, 0, img_label_w)
        # �����
        part_id_cell = QTableWidgetItem()
        part_id_cell.setData(Qt.DisplayRole, p.part_id)
        self.materialTableWidget.setItem(i, 1, part_id_cell)
        # �ѷ������ϵ���ϸ��Ϣ
        a_dict = self.__parent.detail_assigned_dict if self.__mode == 0 else {}
        part_a_detail_list = a_dict[p.get_part_id()] if p.get_part_id() in a_dict else None
        # ����
        t_des_data = [p.name]
        if p.description is not None and p.description != '':
            t_des_data.append(p.description)
        b = p.get_specified_tag(self.__database, 'Ʒ��')
        s = p.get_specified_tag(self.__database, '��׼')
        if b is not None and b != '':
            t_des_data.append(b)
        if s is not None and s != '':
            t_des_data.append(s)
        des_cell = QTableWidgetItem()
        des_cell.setData(Qt.DisplayRole, '\n'.join(t_des_data))
        self.materialTableWidget.setItem(i, 2, des_cell)
        # ����
        qty_cell_w = NoWheelPinBox()
        qty_cell_w.setMaximum(9999.0)
        qty_cell_w.setValue(qty)
        qty_cell_w.setAlignment(Qt.AlignRight)
        self.materialTableWidget.setCellWidget(i, 3, qty_cell_w)
        if self.__mode == 0:
            qty_cell_w.valueChanged.connect(self.__change_qty)
            self.__spin_box_dict[qty_cell_w] = part_id_cell
        # ��λ
        u = p.get_specified_tag(self.__database, '��λ')
        if u != '':
            unit_cell = QTableWidgetItem()
            unit_cell.setData(Qt.DisplayRole, u)
            unit_cell.setTextAlignment(Qt.AlignCenter)
            self.materialTableWidget.setItem(i, 4, unit_cell)
        # �ִ���Ϣ
        qty_in_storing = 0.  # �ֿ��������
        qty_in_site = 0.  # �ֳ�������
        storing_data_s = self.__database.get_storing(part_id=p.get_part_id())
        zd_erp_id = p.get_specified_tag(self.__database, '�����е�ERP���ϱ���')
        jo_erp_id = p.get_specified_tag(self.__database, '��ŷERP���ϱ���')
        remain_qty = qty  # ʣ�������
        # �ѷ��������
        s1 = 0.0
        s2 = 0.0
        s3 = 0.0
        a_dict = self.__parent.assigned_material_dict if self.__mode == 0 else self.__local_assigned_data
        if (a_dict is not None) and (p.get_part_id() in a_dict):
            ss = a_dict[p.get_part_id()]
            s1 = ss[0]  # ��ŷ�ֿ�
            s2 = ss[1]  # �е²ֿ�
            s3 = ss[2]  # �ֳ�
        # �ֳ��ִ���Ϣ
        if storing_data_s is not None:
            s_qty = 0.
            u_date = None
            for r_s in storing_data_s:
                if r_s[1] != 'A':
                    continue
                s_qty += r_s[2]
                if (u_date is not None and r_s[3] > u_date) or u_date is None:
                    u_date = r_s[3]
            s_o_qty = s_qty
            s_qty -= s3
            qty_in_site += s_qty
            if u_date is not None and s_o_qty > 0.:
                if pick_datas is not None:
                    vv = pick_datas[2]
                else:
                    if s_qty >= remain_qty:
                        vv = remain_qty
                    else:
                        vv = s_qty
                    remain_qty -= vv
                qty_w = QtyComWidget(s_qty + s3, vv, parent=self.materialTableWidget, max_value=s_qty, a_qty=s3)
                if part_a_detail_list is not None:
                    qty_w.set_assigned_detail_data(part_a_detail_list, 3)
                self.materialTableWidget.setCellWidget(i, 11, qty_w)
                u_date_cell = QTableWidgetItem()
                if type(u_date) == str:
                    u_date = datetime.datetime.fromisoformat(u_date)
                u_date_cell.setData(Qt.DisplayRole, u_date.strftime('%Y/%m/%d'))
                self.materialTableWidget.setItem(i, 12, u_date_cell)
        # ��ŷ�ִ���Ϣ
        if jo_erp_id != '':
            erp_id_cell = QTableWidgetItem()
            erp_id_cell.setData(Qt.DisplayRole, jl_erp_id)
            self.materialTableWidget.setItem(i, 8, erp_id_cell)
            s_qty = 0.
            u_date = None
            if storing_data_s is not None:
                for r_s in storing_data_s:
                    if r_s[1] != 'J':
                        continue
                    s_qty += r_s[2]
                    if (u_date is not None and r_s[3] > u_date) or u_date is None:
                        u_date = r_s[3]
                s_o_qty = s_qty
                s_qty -= s2
                qty_in_storing += s_qty
                if u_date is not None and s_o_qty > 0.:
                    if pick_datas is not None:
                        vv = pick_datas[1]
                    else:
                        if s_qty >= remain_qty:
                            vv = remain_qty
                        else:
                            vv = s_qty
                        remain_qty -= vv
                    qty_w = QtyComWidget(s_qty + s2, vv, parent=self.materialTableWidget, max_value=s_qty,
                                         a_qty=s2)
                    if part_a_detail_list is not None:
                        qty_w.set_assigned_detail_data(part_a_detail_list, 2)
                    self.materialTableWidget.setCellWidget(i, 9, qty_w)
                    u_date_cell = QTableWidgetItem()
                    if type(u_date) == str:
                        u_date = datetime.datetime.fromisoformat(u_date)
                    u_date_cell.setData(Qt.DisplayRole, u_date.strftime('%Y/%m/%d'))
                    self.materialTableWidget.setItem(i, 10, u_date_cell)
        # �е�ERP�ִ���Ϣ
        if zd_erp_id != '':
            erp_id_cell = QTableWidgetItem()
            erp_id_cell.setData(Qt.DisplayRole, zd_erp_id)
            self.materialTableWidget.setItem(i, 5, erp_id_cell)
            s_qty = 0.
            u_date = None
            if storing_data_s is not None:
                for r_s in storing_data_s:
                    if r_s[1] != 'D' and r_s[1] != 'E':
                        continue
                    s_qty += r_s[2]
                    if (u_date is not None and r_s[3] > u_date) or u_date is None:
                        u_date = r_s[3]
                s_o_qty = s_qty
                s_qty -= s1
                qty_in_storing += s_qty
                if u_date is not None and s_o_qty > 0.:
                    if pick_datas is not None:
                        vv = pick_datas[0]
                    else:
                        if s_qty >= remain_qty:
                            vv = remain_qty
                        else:
                            vv = s_qty
                        remain_qty -= vv
                    qty_w = QtyComWidget(s_qty + s1, vv, parent=self.materialTableWidget, max_value=s_qty, a_qty=s1)
                    if part_a_detail_list is not None:
                        qty_w.set_assigned_detail_data(part_a_detail_list, 1)
                    self.materialTableWidget.setCellWidget(i, 6, qty_w)
                    u_date_cell = QTableWidgetItem()
                    if type(u_date) == str:
                        u_date = datetime.datetime.fromisoformat(u_date)
                    u_date_cell.setData(Qt.DisplayRole, u_date.strftime('%Y/%m/%d'))
                    self.materialTableWidget.setItem(i, 7, u_date_cell)
        # ���ϲ���
        supply_type = p.get_specified_tag(self.__database, '���ϲ���')
        if supply_type != '':
            supply_type_cell = QTableWidgetItem()
            supply_type_cell.setData(Qt.DisplayRole, supply_type)
            self.materialTableWidget.setItem(i, 13, supply_type_cell)
        # �ɹ��ƻ�
        plan_p: dict = self.__parent.purchase_plan
        if plan_p is not None:
            if zd_erp_id in plan_p:
                pd = plan_p[zd_erp_id]
                pd_i = []
                for pd_j in pd:
                    pd_i.append(f'{pd_j[0]} {pd_j[1]:.2f} {pd_j[2]:.2f}')
                plan_cell = QTableWidgetItem()
                plan_cell.setData(Qt.DisplayRole, '\n'.join(pd_i))
                self.materialTableWidget.setItem(i, 14, plan_cell)
        # ���������Աȣ��ı��е���ɫ
        if qty > qty_in_storing + qty_in_site:
            part_id_cell.setBackground(QColor(255, 0, 0))
        # ���ڴ洢���ݵĵ�Ԫ
        part_id_cell.setData(Qt.UserRole, (p, qty_in_storing, qty_in_site, qty))
        self.materialTableWidget.setRowHeight(i, 66)

    def __fill_one_row(self, p: Part, qty, row_index, pick_datas=None):
        """
        ����һ��
        :param p:
        :param qty:
        :param row_index:
        :param pick_datas: (�е��ѷ��䣬 �����ѷ��䣬 �ֳ��ѷ���)
        :return:
        """
        i = row_index
        original_item = self.materialTableWidget.item(row_index, 2)
        if original_item is not None:
            # ɾ��ԭ����Ϣ
            self.materialTableWidget.insertRow(i)
            self.materialTableWidget.removeRow(i + 1)
        # ͼʾ
        img_data = self.__database.get_thumbnail_2_part(p.get_part_id())
        if img_data is not None:
            img = QPixmap()
            img.loadFromData(img_data)
            n_img = img.scaled(64, 64, aspectRatioMode=Qt.KeepAspectRatio)
            img_label_w = QLabel()
            img_label_w.setPixmap(n_img)
            img_label_w.setAlignment(Qt.AlignCenter)
            self.materialTableWidget.setCellWidget(i, 0, img_label_w)
        # �����
        part_id_cell = QTableWidgetItem()
        part_id_cell.setData(Qt.DisplayRole, p.part_id)
        self.materialTableWidget.setItem(i, 1, part_id_cell)
        # �ѷ������ϵ���ϸ��Ϣ
        a_dict = self.__parent.detail_assigned_dict if self.__mode == 0 else {}
        part_a_detail_list = a_dict[p.get_part_id()] if p.get_part_id() in a_dict else None
        # ����
        t_des_data = [p.name]
        if p.description is not None and p.description != '':
            t_des_data.append(p.description)
        b = p.get_specified_tag(self.__database, 'Ʒ��')
        s = p.get_specified_tag(self.__database, '��׼')
        if b is not None and b != '':
            t_des_data.append(b)
        if s is not None and s != '':
            t_des_data.append(s)
        des_cell = QTableWidgetItem()
        des_cell.setData(Qt.DisplayRole, '\n'.join(t_des_data))
        self.materialTableWidget.setItem(i, 2, des_cell)
        # ����
        qty_cell_w = NoWheelPinBox()
        qty_cell_w.setMaximum(9999.0)
        qty_cell_w.setValue(qty)
        qty_cell_w.setAlignment(Qt.AlignRight)
        self.materialTableWidget.setCellWidget(i, 3, qty_cell_w)
        if self.__mode == 0:
            qty_cell_w.valueChanged.connect(self.__change_qty)
            self.__spin_box_dict[qty_cell_w] = part_id_cell
        # ��λ
        u = p.get_specified_tag(self.__database, '��λ')
        if u != '':
            unit_cell = QTableWidgetItem()
            unit_cell.setData(Qt.DisplayRole, u)
            unit_cell.setTextAlignment(Qt.AlignCenter)
            self.materialTableWidget.setItem(i, 4, unit_cell)
        # �ִ���Ϣ
        qty_in_storing = 0.  # �ֿ��������
        qty_in_site = 0.  # �ֳ�������
        storing_data_s = self.__database.get_storing(part_id=p.get_part_id())
        zd_erp_id = p.get_specified_tag(self.__database, '�����е�ERP���ϱ���')
        jl_erp_id = p.get_specified_tag(self.__database, '��������ERP���ϱ���')
        remain_qty = qty  # ʣ�������
        # �ѷ��������
        s1 = 0.0
        s2 = 0.0
        s3 = 0.0
        a_dict = self.__parent.assigned_material_dict if self.__mode == 0 else self.__local_assigned_data
        if (a_dict is not None) and (p.get_part_id() in a_dict):
            ss = a_dict[p.get_part_id()]
            s1 = ss[0]  # �е²ֿ�
            s2 = ss[1]  # ���ֲֿ�
            s3 = ss[2]  # �ֳ�
        # �ֳ��ִ���Ϣ
        if storing_data_s is not None:
            s_qty = 0.
            u_date = None
            for r_s in storing_data_s:
                if r_s[1] != 'A':
                    continue
                s_qty += r_s[2]
                if (u_date is not None and r_s[3] > u_date) or u_date is None:
                    u_date = r_s[3]
            s_o_qty = s_qty
            s_qty -= s3
            qty_in_site += s_qty
            if u_date is not None and s_o_qty > 0.:
                if pick_datas is not None:
                    vv = pick_datas[2]
                else:
                    if s_qty >= remain_qty:
                        vv = remain_qty
                    else:
                        vv = s_qty
                    remain_qty -= vv
                qty_w = QtyComWidget(s_qty + s3, vv, parent=self.materialTableWidget, max_value=s_qty, a_qty=s3)
                if part_a_detail_list is not None:
                    qty_w.set_assigned_detail_data(part_a_detail_list, 3)
                self.materialTableWidget.setCellWidget(i, 11, qty_w)
                u_date_cell = QTableWidgetItem()
                if type(u_date) == str:
                    u_date = datetime.datetime.fromisoformat(u_date)
                u_date_cell.setData(Qt.DisplayRole, u_date.strftime('%Y/%m/%d'))
                self.materialTableWidget.setItem(i, 12, u_date_cell)
        # �е�ERP�ִ���Ϣ
        if zd_erp_id != '':
            erp_id_cell = QTableWidgetItem()
            erp_id_cell.setData(Qt.DisplayRole, zd_erp_id)
            self.materialTableWidget.setItem(i, 5, erp_id_cell)
            s_qty = 0.
            u_date = None
            if storing_data_s is not None:
                for r_s in storing_data_s:
                    if r_s[1] != 'D' and r_s[1] != 'E':
                        continue
                    s_qty += r_s[2]
                    if (u_date is not None and r_s[3] > u_date) or u_date is None:
                        u_date = r_s[3]
                s_o_qty = s_qty
                s_qty -= s1
                qty_in_storing += s_qty
                if u_date is not None and s_o_qty > 0.:
                    if pick_datas is not None:
                        vv = pick_datas[0]
                    else:
                        if s_qty >= remain_qty:
                            vv = remain_qty
                        else:
                            vv = s_qty
                        remain_qty -= vv
                    qty_w = QtyComWidget(s_qty + s1, vv, parent=self.materialTableWidget, max_value=s_qty, a_qty=s1)
                    if part_a_detail_list is not None:
                        qty_w.set_assigned_detail_data(part_a_detail_list, 1)
                    self.materialTableWidget.setCellWidget(i, 6, qty_w)
                    u_date_cell = QTableWidgetItem()
                    if type(u_date) == str:
                        u_date = datetime.datetime.fromisoformat(u_date)
                    u_date_cell.setData(Qt.DisplayRole, u_date.strftime('%Y/%m/%d'))
                    self.materialTableWidget.setItem(i, 7, u_date_cell)
        # ����ERP�ִ���Ϣ
        if jl_erp_id != '':
            erp_id_cell = QTableWidgetItem()
            erp_id_cell.setData(Qt.DisplayRole, jl_erp_id)
            self.materialTableWidget.setItem(i, 8, erp_id_cell)
            s_qty = 0.
            u_date = None
            if storing_data_s is not None:
                for r_s in storing_data_s:
                    if r_s[1] != 'F':
                        continue
                    s_qty += r_s[2]
                    if (u_date is not None and r_s[3] > u_date) or u_date is None:
                        u_date = r_s[3]
                s_o_qty = s_qty
                s_qty -= s2
                qty_in_storing += s_qty
                if u_date is not None and s_o_qty > 0.:
                    if pick_datas is not None:
                        vv = pick_datas[1]
                    else:
                        if s_qty >= remain_qty:
                            vv = remain_qty
                        else:
                            vv = s_qty
                        remain_qty -= vv
                    qty_w = QtyComWidget(s_qty + s2, vv, parent=self.materialTableWidget, max_value=s_qty, a_qty=s2)
                    if part_a_detail_list is not None:
                        qty_w.set_assigned_detail_data(part_a_detail_list, 2)
                    self.materialTableWidget.setCellWidget(i, 9, qty_w)
                    u_date_cell = QTableWidgetItem()
                    if type(u_date) == str:
                        u_date = datetime.datetime.fromisoformat(u_date)
                    u_date_cell.setData(Qt.DisplayRole, u_date.strftime('%Y/%m/%d'))
                    self.materialTableWidget.setItem(i, 10, u_date_cell)
        # ���ϲ���
        supply_type = p.get_specified_tag(self.__database, '���ϲ���')
        if supply_type != '':
            supply_type_cell = QTableWidgetItem()
            supply_type_cell.setData(Qt.DisplayRole, supply_type)
            self.materialTableWidget.setItem(i, 13, supply_type_cell)
        # �ɹ��ƻ�
        plan_p: dict = self.__parent.purchase_plan
        if plan_p is not None:
            if zd_erp_id in plan_p:
                pd = plan_p[zd_erp_id]
                pd_i = []
                for pd_j in pd:
                    pd_i.append(f'{pd_j[0]} {pd_j[1]:.2f} {pd_j[2]:.2f}')
                plan_cell = QTableWidgetItem()
                plan_cell.setData(Qt.DisplayRole, '\n'.join(pd_i))
                self.materialTableWidget.setItem(i, 14, plan_cell)
        # ���������Աȣ��ı��е���ɫ
        if qty > qty_in_storing + qty_in_site:
            part_id_cell.setBackground(QColor(255, 0, 0))
        # ���ڴ洢���ݵĵ�Ԫ
        part_id_cell.setData(Qt.UserRole, (p, qty_in_storing, qty_in_site, qty))
        self.materialTableWidget.setRowHeight(i, 66)

    def get_all_material(self):
        """
        ��ȡ��ǰ���嵥
        :return: [�������������ϣ����ֳ����ϵ������嵥�����е²ֿ����ϵ������嵥���Ӿ��ֲֿ����ϵ������嵥����������]
        """
        pick_from_site = []  # ���ֳ������� (part, erp_id, qty)
        pick_from_zd_storage = []  # ���е²ֿ������
        pick_from_jl_storage = []  # �þ��ֻ����˲ֿ������
        all_material = []  # �������� (part, qty, available_qty)
        qty_not_enough = 0
        row_c = self.materialTableWidget.rowCount()
        if row_c < 1:
            raise Exception('û�в����嵥��')
        for i in range(row_c):
            item = self.materialTableWidget.item(i, 1)
            dd = item.data(Qt.UserRole)
            available_qty = dd[1] + dd[2]
            if available_qty < dd[3]:
                qty_not_enough += 1
            all_material.append((dd[0], dd[3], available_qty))
            w = self.materialTableWidget.cellWidget(i, 11)
            if w is not None and w.value() > 0.:
                pick_from_site.append((dd[0], '', w.value()))
            w = self.materialTableWidget.cellWidget(i, 6)
            if w is not None and w.value() > 0.:
                zd_erp_id = self.materialTableWidget.item(i, 5).data(Qt.DisplayRole)
                pick_from_zd_storage.append((dd[0], zd_erp_id, w.value()))
            w = self.materialTableWidget.cellWidget(i, 9)
            if w is not None and w.value() > 0.:
                jl_erp_id = self.materialTableWidget.item(i, 8).data(Qt.DisplayRole)
                pick_from_jl_storage.append((dd[0], jl_erp_id, w.value()))
        return qty_not_enough, pick_from_site, pick_from_zd_storage, pick_from_jl_storage, all_material

    def fill_data(self, parts_dict):
        if self.__mode == 0:
            p: MaterialTab = self.__parent
            p.get_assigned_materials()
        self.materialTableWidget.clear()
        self.set_header()
        parts = list(parts_dict.keys())
        r_c = len(parts)
        self.materialTableWidget.setRowCount(r_c)
        i = 0
        for r in parts:
            p: Part = r
            qty = parts_dict[p]
            self.__fill_one_row(p, qty, i)
            i += 1
        QTableWidget.resizeColumnsToContents(self.materialTableWidget)
        # QTableWidget.resizeRowsToContents( self.materialTableWidget )


class PartSelectDialog(QDialog, BlankDialog):

    def __init__(self, parent, database, _data, assigned_data=None, purchase_plan=None):
        super(PartSelectDialog, self).__init__(parent)
        self.__database = database
        self.__data = _data
        self.__assigned_data = assigned_data
        self.purchase_plan = purchase_plan
        self.selected_data = None
        self.selected_table = MaterialTablePanel(self, database, mode=1, assigned_data=assigned_data)
        self.setup_ui()
        self.data_init(_data)

    def setup_ui(self):
        super(PartSelectDialog, self).setupUi(self)
        v_layout = QVBoxLayout()
        v_layout.addWidget(self.selected_table)
        v_layout.addWidget(self.buttonBox)
        self.setLayout(v_layout)
        self.setMinimumWidth(800)
        self.setMinimumHeight(400)

    def data_init(self, the_data):
        self.selected_table.fill_data(the_data)

    def accept(self):
        the_table: QTableWidget = self.selected_table.materialTableWidget
        c_r = the_table.currentRow()
        if c_r < 0:
            QMessageBox.warning(self, '', 'û��ѡ��')
            return
        d_item = the_table.item(c_r, 1)
        d_w: NoWheelPinBox = the_table.cellWidget(c_r, 3)
        qty = d_w.value()
        d_data = d_item.data(Qt.UserRole)
        self.selected_data = (d_data[0], qty)
        self.done(QDialog.Accepted)


class GetPurchasePlan(QThread):
    finish_signal = pyqtSignal(int, str)  # ����ź�

    def __init__(self, _dir):
        """
        �����ɹ��ƻ����߳�
        :param _dir: �ɹ��ƻ��ļ����õ�Ŀ¼
        """
        super(GetPurchasePlan, self).__init__()
        self.__dir = _dir
        self.purchase_plan = {}

    def run(self) -> None:
        # ��ȡ�ɹ���������
        try:
            self.purchase_plan.clear()
            files = os.listdir(self.__dir)
            for f in files:
                dot_index = f.rindex('.')
                file_type = f[dot_index:]
                if file_type.upper() != '.XLSX' or f[0] == '_' or f[0] == '~':
                    continue
                full_path = os.path.join(self.__dir, f)
                wb = load_workbook(full_path, read_only=True, data_only=True)
                ws_s = wb.sheetnames
                bill_code_dict = {}
                # �ռ����ݣ�ȥ���ɹ��ƻ��У��ظ�������
                for s in ws_s:
                    ws = wb[s]
                    r_i = 0
                    for r in ws.iter_rows():
                        if r_i == 0:
                            r_i += 1
                            continue
                        erp_id = r[3].value
                        if erp_id is None:
                            break
                        require_bill = r[2].value
                        bill_code = f'{require_bill} {erp_id}'
                        d1 = r[10].value
                        req_qty = float(d1) if d1 is not None else 0.0
                        d2 = r[12].value
                        pur_qty = float(d2) if d2 is not None else 0.0
                        d3 = r[13].value
                        income_qty = float(d3) if d3 is not None else 0.0
                        pb = r[6].value
                        produce_bill = '' if pb is None else pb.strip()
                        if bill_code in bill_code_dict:
                            # ���������з����������
                            o_data = bill_code_dict[bill_code]
                            bill_code_dict[bill_code] = (produce_bill, req_qty, pur_qty, income_qty + o_data[3])
                        else:
                            bill_code_dict[bill_code] = (produce_bill, req_qty, pur_qty, income_qty)
                        r_i += 1
                # ��������
                for k in bill_code_dict.keys():
                    d_data = bill_code_dict[k]
                    erp_id = k[-13:]
                    produce_bill = d_data[0]
                    req_qty = d_data[1]
                    pur_qty = d_data[2]
                    income_qty = d_data[3]
                    if income_qty >= req_qty and income_qty >= pur_qty:
                        continue
                    need_pur_qty = req_qty - max(pur_qty, income_qty)
                    pur_qty = pur_qty - income_qty
                    if erp_id in self.purchase_plan:
                        ll = self.purchase_plan[erp_id]
                        if produce_bill in ll:
                            dd = ll[produce_bill]
                            ll[produce_bill] = (dd[0] + need_pur_qty, dd[1] + pur_qty)
                        else:
                            ll[produce_bill] = (need_pur_qty, pur_qty)
                    else:
                        self.purchase_plan[erp_id] = {produce_bill: (need_pur_qty, pur_qty)}
                for k in self.purchase_plan.keys():
                    dd = self.purchase_plan[k]
                    nd = []
                    for d in dd.keys():
                        ndd = dd[d]
                        nd.append((d, ndd[0], ndd[1]))
                    self.purchase_plan[k] = nd
            self.finish_signal.emit(len(self.purchase_plan), 'Done.')
        except Exception as ex:
            self.finish_signal.emit(-1, f'�����ɹ��ƻ�ʱ�쳣 -> {ex.__str__()}')


# noinspection PyUnresolvedReferences
class MaterialTab(QWidget):

    def __init__(self, parent, database, **other):
        super(MaterialTab, self).__init__(parent)
        self.__dir = other['_dir'] if '_dir' in other else None
        self.__database = database
        self.__parent = parent
        # �ڿ��ÿ���У��ѷ�����嵥�ļ��б�
        self.assigned_list_widget = QListWidget(self)
        # ��ǰ���ڼƻ����ϵ��ļ�
        self.current_material_4_pick_up = None
        # �嵥�ļ��б���Ҽ��˵�
        self.__menu_4_list = QMenu(parent=self.assigned_list_widget)
        self.__pick_up_action = self.__menu_4_list.addAction('����')
        self.__summary_action = self.__menu_4_list.addAction('�ۼ�')
        # �ѷ�����ϵ����� part_id : [�ѷ�����е�����, �ѷ���ľ�������, �ѷ�����ֳ�����]
        self.assigned_material_dict = other['assigned_data'] if 'assigned_data' in other else {}
        # �е����ϲɹ��ƻ��ļ�
        self.purchase_plan_dir = other['plan'] if 'plan' in other else None
        # �ɹ��ƻ� -> {ERP���ϱ��룺[��ͬ�ţ�δ�ɹ������ɹ���]}
        self.purchase_plan = {}
        # ��ȡ�ɹ��ƻ���ɵı�־
        self.purchase_plan_done = False
        # ������������ֵ�
        self.detail_assigned_dict = {}
        # ���вɹ��ƻ��������߳�
        self.__a_thread = GetPurchasePlan(self.purchase_plan_dir)
        # ���Ͻǵ�������ť
        self.add_tab_btn = QPushButton('+')
        self.remove_tab_btn = QPushButton('-')
        self.save_table_btn = QPushButton('����')
        self.__tab_widget = QTabWidget(self)
        self.__menu_4_tab = QMenu(self.__tab_widget)
        self.rename_action = self.__menu_4_tab.addAction('������')
        self.__w_list = []
        # ��ȡ�ѷ�������ʱ�����Ե��ļ�����������
        self.ignored_assigned_files = None
        # ��ʱֹͣĳЩ��Ӧ����
        self.__pause_tab_change_handler = False
        self.setup_ui()
        self.__initialization = True
        self.init_datas()
        self.__initialization = False

    def setup_ui(self):
        top_widget = QWidget()
        top_h_layout = QHBoxLayout()
        top_h_layout.addWidget(QLabel('�ѷ�������嵥�ļ�'))
        top_h_layout.addWidget(self.assigned_list_widget)
        # self.assigned_list_widget.setMinimumHeight(60)
        top_widget.setLayout(top_h_layout)
        v_splitter = QSplitter(Qt.Vertical, self)
        v_splitter.addWidget(top_widget)
        v_splitter.addWidget(self.__tab_widget)
        v_splitter.setStretchFactor(0, 1)
        v_splitter.setStretchFactor(1, 5)
        h_layout = QHBoxLayout(self)
        h_layout.addWidget(v_splitter)
        self.setLayout(h_layout)
        self.assigned_list_widget.setSelectionMode(QAbstractItemView.SingleSelection)
        self.assigned_list_widget.doubleClicked.connect(self.__open_assigned_file)
        self.assigned_list_widget.setContextMenuPolicy(Qt.CustomContextMenu)
        self.assigned_list_widget.customContextMenuRequested.connect(self.__on_custom_context_menu_requested)
        self.__summary_action.triggered.connect(self.__do_material_summary)
        self.__pick_up_action.triggered.connect(self.__pick_up_assigned_material)
        self.__a_thread.finish_signal.connect(self.analysis_purchase_plan)

        # tab���ӹ���
        w = QWidget()
        w_layout = QHBoxLayout()
        w_layout.addWidget(self.add_tab_btn)
        w_layout.addWidget(self.remove_tab_btn)
        w_layout.addWidget(self.save_table_btn)
        w.setLayout(w_layout)
        w.setFixedWidth(120)
        w.setFixedHeight(40)
        self.__tab_widget.setCornerWidget(w)
        self.add_tab_btn.clicked.connect(self.__add_tab)
        self.remove_tab_btn.clicked.connect(self.__remove_tab)
        self.__tab_widget.setContextMenuPolicy(Qt.CustomContextMenu)
        self.__tab_widget.customContextMenuRequested.connect(self.__on_custom_context_menu_requested)
        self.rename_action.triggered.connect(self.__rename_tab)
        self.__tab_widget.currentChanged.connect(self.__tab_changed)
        self.save_table_btn.clicked.connect(self.update_assigned_file)

    def update_assigned_file(self):
        # ���µ�ǰ���ѷ������ϵ��ļ�
        current_index = self.__tab_widget.currentIndex()
        sheet_name = self.__tab_widget.tabText(current_index)
        file_name = self.current_material_4_pick_up
        material_table: MaterialTablePanel = self.__w_list[current_index]
        try:
            if is_used(file_name):
                raise Exception(f'{file_name}�ļ���ռ�ã��޷�������')
            wb = load_workbook(file_name)
            ws = wb[sheet_name]
            y = 2
            part_id_row_dict = {}
            while True:
                part_id_cell = ws.cell(row=y, column=1)
                part_id = part_id_cell.value
                if part_id is None:
                    break
                part_id_row_dict[part_id] = y
                y += 1
            all_materials = material_table.get_all_material()
            if len(all_materials[4]) < 1:
                raise Exception('û�������嵥')
            zd_storage_dict = dict()
            for m in all_materials[2]:
                zd_storage_dict[m[0]] = m[2]
            jl_storage_dict = dict()
            for m in all_materials[3]:
                jl_storage_dict[m[0]] = m[2]
            site_storage_dict = dict()
            for m in all_materials[1]:
                site_storage_dict[m[0]] = m[2]
            for r in all_materials[4]:
                p: Part = r[0]
                qty = r[1]
                qty_zd = zd_storage_dict[p] if p in zd_storage_dict.keys() else 0.0
                qty_jl = jl_storage_dict[p] if p in jl_storage_dict.keys() else 0.0
                qty_site = site_storage_dict[p] if p in site_storage_dict.keys() else 0.0
                y = part_id_row_dict[p.get_part_id()]
                qty_cell = ws.cell(row=y, column=7)
                p1 = ws.cell(row=y, column=15).value
                p2 = ws.cell(row=y, column=16).value
                p3 = ws.cell(row=y, column=17).value
                ws.cell(row=y, column=9).value = qty_zd
                ws.cell(row=y, column=10).value = qty_jl
                ws.cell(row=y, column=11).value = qty_site
                qty_cell.value = qty + p1 + p2 + p3
            wb.save(file_name)
            material_table.table_need_save = False
            QMessageBox.information(None, '', '������ɡ�')
        except Exception as ex:
            QMessageBox.warning(self, '', ex.__str__(), QMessageBox.Yes)

    def set_save_btn(self, is_enable):
        # ���ñ��水ť�Ŀ���
        self.save_table_btn.setEnabled(is_enable)

    def __tab_changed(self, _index):
        if self.__pause_tab_change_handler or self.__initialization:
            return
        # ��ǩҳ�ı����Ӧ����
        no_tab = self.__tab_widget.count() < 1
        if not no_tab:
            current_tab_index = self.__tab_widget.currentIndex()
            m_table: MaterialTablePanel = self.__w_list[current_tab_index]
            self.save_table_btn.setEnabled(m_table.table_need_save)
        else:
            self.save_table_btn.setEnabled(False)
        self.remove_tab_btn.setEnabled(not no_tab)

    def __rename_tab(self):
        """
        ������һ����ǩҳ
        :return:
        """
        current_tab_index = self.__tab_widget.currentIndex()
        _name = self.__tab_widget.tabText(current_tab_index)
        tab_name, ok = QInputDialog.getText(self.__parent, '����', 'ҳ��������', QLineEdit.Normal, _name)
        if not ok:
            return
        self.__tab_widget.setTabText(current_tab_index, tab_name)

    def __add_tab(self):
        """
        ����һ����ǩҳ
        :return:
        """
        tab_count = len(self.__w_list) + 1
        tab_name, ok = QInputDialog.getText(self.__parent, '����', '��ҳ������', QLineEdit.Normal, f'ҳ{tab_count}')
        if not ok:
            return
        neu_tab = MaterialTablePanel(self, self.__database, mode=0)
        self.__w_list.append(neu_tab)
        self.__tab_widget.addTab(neu_tab, tab_name)

    def __remove_tab(self):
        """
        �Ƴ�һ��Tab
        :return:
        """
        current_tab_index = self.__tab_widget.currentIndex()
        if current_tab_index >= 0:
            self.__tab_widget.removeTab(current_tab_index)
            w = self.__w_list[current_tab_index]
            self.__w_list.remove(w)

    def get_current_material_table(self) -> MaterialTablePanel:
        """
        ��ȡ��ǰ����������ݱ��
        :return:
        """
        current_tab = self.__tab_widget.currentWidget()
        tab_text = self.__tab_widget.tabText(self.__tab_widget.currentIndex())
        return current_tab, tab_text

    def init_datas(self):
        # ��ʼ������
        try:
            self.__analysis_file(self.__dir)
            self.get_assigned_materials()
            self.__w_list.append(MaterialTablePanel(self, self.__database, mode=0))
            self.__tab_widget.addTab(self.__w_list[0], '�����嵥')
            self.__a_thread.start()
        except Exception as ex:
            QMessageBox.warning(self, '', f'��ʼ��ʱ�쳣��{ex.__str__()}')
            # raise ex

    def analysis_purchase_plan(self, count, message):
        # ��ȡ�ɹ���������
        if count < 0:
            self.__parent.show_status_message(message)
        else:
            self.purchase_plan = self.__a_thread.purchase_plan
            self.__parent.show_status_message(f'�ɹ��ƻ���ȡ��ɣ�����ȡ{count}����Ч��¼��')
        self.purchase_plan_done = True

    def __on_custom_context_menu_requested(self, pos: QPoint):
        if self.sender() is self.assigned_list_widget:
            on_item = self.assigned_list_widget.itemAt(pos)
            self.__pick_up_action.setVisible(on_item is not None)
            self.__menu_4_list.exec_(QCursor.pos())
        elif self.sender() is self.__tab_widget:
            current_tab_index = self.__tab_widget.currentIndex()
            if current_tab_index >= 0:
                self.__menu_4_tab.exec_(QCursor.pos())

    def show_status_message(self, msg, sec=None):
        self.__parent.show_status_message(msg, sec)

    def __open_assigned_file(self):
        # ���ļ�
        the_item = self.assigned_list_widget.currentItem()
        if the_item is None:
            return
        os.startfile(the_item.text())

    def __do_material_summary(self):
        resp = QMessageBox.question(self, '', 'ȷ��Ҫ�����ۼƣ�', QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if resp == QMessageBox.No:
            return
        # �ۼ��ѷ�������ϣ������ΪExcel�ļ�
        sum_dict = {}  # �ܵ��ۼ��嵥
        product_sum_dict = {}  # ���ڲ�Ʒ�������ֲ��嵥
        product_column_list = []  # ��Ʒ�����ֲ��嵥���к�
        wwb = xw.Book()
        wws = wwb.sheets.active
        wws.name = 'summary'
        c = self.assigned_list_widget.count()
        # �����б�ͷ
        headers = [
            '�����', '����', '�ͺ�����', 'Ʒ��', '��׼', '��λ', '����', '�����е����ϱ���', '�е²ֿ���',
            '���ֲֿ���', '�ֳ����', '����', '���ϲ���', '��治��', 'δ�ɹ�', '�ɹ�δ����', '�ɹ��ƻ�']
        for i in range(c):
            current_item = self.assigned_list_widget.item(i)
            f_name = current_item.text()
            product_name = os.path.basename(f_name)[:-5]
            product_column_list.append(product_name)
        headers.extend(product_column_list)
        wws.range((1, 1), (1, len(headers) + 1)).value = headers
        # ��ȡ����
        for i in range(c):
            current_item = self.assigned_list_widget.item(i)
            f_name = current_item.text()
            product_name = os.path.basename(f_name)[:-5]
            s_dict = {}
            try:
                wb = load_workbook(f_name, data_only=True)
                ws_s = wb.sheetnames
                for s in ws_s:
                    ws = wb[s]
                    r_i = 0
                    for r in ws.iter_rows():
                        if r_i == 0:
                            r_i += 1
                            continue
                        p_id_cell = r[0].value
                        n_qty = r[6].value
                        p1 = r[14].value
                        p2 = r[15].value
                        p3 = r[16].value
                        qty = n_qty - p1 - p2 - p3  # ��ȡʣ�໹û���ϵ�����
                        if qty <= 0.0:
                            continue
                        s1 = float(r[8].value)
                        s2 = float(r[9].value)
                        s3 = float(r[10].value)
                        if p_id_cell in s_dict:
                            ss = s_dict[p_id_cell]
                            ss[0] += qty
                            ss[1] += s1
                            ss[2] += s2
                            ss[3] += s3
                        else:
                            s_dict[p_id_cell] = [qty, s1, s2, s3]
                        r_i += 1
                product_sum_dict[product_name] = s_dict
            except Exception as ex:
                self.__parent.show_status_message(f'{f_name} -> ����{ex.__str__()}')
                current_item.setBackground(QColor(255, 0, 0))
        for k in product_sum_dict.keys():
            dd = product_sum_dict[k]
            for part_id in dd.keys():
                part_qty = dd[part_id]
                if part_id in sum_dict:
                    ss = sum_dict[part_id]
                    ss[0] += part_qty[0]
                    ss[1] += part_qty[1]
                    ss[2] += part_qty[2]
                    ss[3] += part_qty[3]
                    ss[4].append(f'{k};{part_qty[0]}')
                else:
                    ss = [part_qty[0], part_qty[1], part_qty[2], part_qty[3], [f'{k};{part_qty[0]}']]
                    sum_dict[part_id] = ss
        # ����嵥
        part_id_list = list(sum_dict.keys())
        part_id_list.sort()
        i = 2
        for part_id in part_id_list:
            dd = sum_dict[part_id]
            p: Part = Part.get_parts(self.__database, part_id=part_id)[0]
            part_id = p.get_part_id()
            id_cell = wws.range((i, 1))
            id_cell.value = part_id
            id_cell.number_format = '00000000'
            name_cell = wws.range((i, 2))
            name_cell.value = p.name
            description_cell = wws.range((i, 3))
            description_cell.value = p.description
            t1 = p.get_specified_tag(self.__database, 'Ʒ��')
            brand_cell = wws.range((i, 4))
            brand_cell.value = t1
            t2 = p.get_specified_tag(self.__database, '��׼')
            standard_cell = wws.range((i, 5))
            standard_cell.value = t2
            t3 = p.get_specified_tag(self.__database, '��λ')
            unit_cell = wws.range((i, 6))
            unit_cell.value = t3
            qty_cell = wws.range((i, 7))
            qty_cell.value = dd[0]
            qty_cell.number_format = '0.00'
            t4 = p.get_specified_tag(self.__database, '�����е�ERP���ϱ���')
            zd_erp_id_cell = wws.range((i, 8))
            zd_erp_id_cell.value = t4
            zd_qty_cell = wws.range((i, 9))
            zd_qty_cell.value = dd[1]
            zd_qty_cell.number_format = '0.00'
            jl_qty_cell = wws.range((i, 10))
            jl_qty_cell.value = dd[2]
            jl_qty_cell.number_format = '0.00'
            site_qty_cell = wws.range((i, 11))
            site_qty_cell.value = dd[3]
            site_qty_cell.number_format = '0.00'
            t5 = p.get_specified_tag(self.__database, '���')
            type_cell = wws.range((i, 12))
            type_cell.value = t5
            t6 = p.get_specified_tag(self.__database, '���ϲ���')
            supply_type_cell = wws.range((i, 13))
            supply_type_cell.value = t6
            warning_cell = wws.range((i, 14))
            warning_cell.formula = f'=SUM(I{i}:K{i})-G{i}'
            # �ɹ��ƻ�����
            if t4 in self.purchase_plan:
                p_dd = self.purchase_plan[t4]
                p_comment = []
                n_p_d = 0.
                p_n_i = 0.
                for p_dd_c in p_dd:
                    n_p_d += p_dd_c[1]
                    p_n_i += p_dd_c[2]
                    p_comment.append(f'{p_dd_c[0]} {p_dd_c[1]:.2f} {p_dd_c[2]:.2f}')
                need_pur_cell = wws.range((i, 15))
                need_pur_cell.value = n_p_d
                need_pur_cell.number_format = '0.00'
                pur_no_income_cell = wws.range((i, 16))
                pur_no_income_cell.value = p_n_i
                pur_no_income_cell.number_format = '0.00'
                pur_comment_cell = wws.range((i, 17))
                pur_comment_cell.value = '\n'.join(p_comment)
            for p_qty in dd[4]:
                tt = p_qty.split(';')
                product_name = tt[0]
                product_qty = float(tt[1])
                c = product_column_list.index(product_name) + 18
                the_cell = wws.range((i, c))
                the_cell.value = product_qty
                the_cell.number_format = '0.00'
            i += 1
        QMessageBox.information(self, '', '�����ϡ�')

    def __analysis_file(self, _dir):
        # �ݹ�ķ����ļ��У�����Excel�ļ����������º��߿�ͷ���ļ���
        try:
            files = os.listdir(_dir)
            for f in files:
                full_path = os.path.join(_dir, f)
                if os.path.isfile(full_path):
                    dot_index = f.rindex('.')
                    file_type = f[dot_index:]
                    if file_type.upper() == '.XLSX' and (not f[0] == '_') and (not f[0] == '~'):
                        neu_item = QListWidgetItem(full_path)
                        neu_item.setData(Qt.UserRole, file_type)
                        self.assigned_list_widget.addItem(neu_item)
                else:
                    if f[0] != '_':
                        self.__analysis_file(full_path)
        except Exception as ex:
            QMessageBox.warning(self, '', f'����{_dir}ʱ������{ex.__str__()}')

    def fill_data(self, material_data, table_index=None) -> int:
        if len(self.__w_list) < 1:
            QMessageBox.warning(None, '', 'û���Ѽ���ı�ǩҳ��')
            return
        self.current_material_4_pick_up = None
        # ���ĳ�����Ĳ�������
        t_index = table_index
        if table_index is None:
            t_index = self.__tab_widget.currentIndex()
        the_table: MaterialTablePanel = self.__w_list[t_index]
        the_table.fill_data(material_data)
        return t_index

    def set_tab_name(self, name, table_index):
        # ����ĳ����ǩҳ������
        self.__tab_widget.setTabText(table_index, name)

    def __pick_up_assigned_material(self):
        self.__pause_tab_change_handler = True
        # ��ȡ�ѷ��������
        try:
            if not self.purchase_plan_done:
                resp = QMessageBox.question(self.__parent, '', '�ɹ��ƻ���δ��ȡ��ɣ�Ҫ������',
                                            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if resp == QMessageBox.No:
                    return
            the_item = self.assigned_list_widget.currentItem()
            self.current_material_4_pick_up = the_item.text()
            self.get_assigned_materials(ignore_files=[the_item.text()])
            wb = load_workbook(the_item.text())
            ws_s = wb.sheetnames
            temp_material_table_list = []
            for s in ws_s:
                a_material_table = MaterialTablePanel(self, self.__database, mode=0)
                a_material_table.in_pick_up_mode = True
                a_material_table.pick_up_list_file = self.current_material_4_pick_up
                ws = wb[s]
                r_i = 0
                assigned_datas = []
                for r in ws.iter_rows():
                    if r_i == 0:
                        r_i += 1
                        continue
                    p_id = r[0].value  # �����
                    if p_id is None:
                        break
                    p_qty = float(r[6].value)  # ������
                    s1 = float(r[8].value)  # �е²ֿ����
                    s2 = float(r[9].value)  # ���ֲֿ����
                    s3 = float(r[10].value)  # �ֳ����Ϸ���
                    p1 = float(r[14].value)  # �е²ֿ�������
                    p2 = float(r[15].value)  # ���ֲֿ�������
                    p3 = float(r[16].value)  # �ֳ�������
                    r_p_qty = p_qty - p1 - p2 - p3
                    if r_p_qty <= 0.0:
                        continue
                    r_s1 = s1 - p1
                    r_s2 = s2 - p2
                    r_s3 = s3 - p3
                    assigned_datas.append((p_id, r_p_qty, r_s1, r_s2, r_s3))
                a_material_table.fill_pick_datas(assigned_datas)
                temp_material_table_list.append((a_material_table, s))
            # ������е�ѡ�
            self.__w_list.clear()
            for t in range(self.__tab_widget.count()):
                self.__tab_widget.removeTab(0)
            for t in temp_material_table_list:
                self.__w_list.append(t[0])
                self.__tab_widget.addTab(t[0], t[1])
        except Exception as ex:
            QMessageBox.warning(self.__parent, '�����ѷ����嵥ʱ�쳣', ex.__str__(), QMessageBox.Yes)
        finally:
            self.__pause_tab_change_handler = False

    def get_assigned_materials(self, ignore_files=None):
        """
        # ��ȡ�ѷ�����ϵ�����
        :param ignore_files: Ҫ���Ե������ļ����б�
        :return:
        """
        if self.__initialization:
            return
        if self.ignored_assigned_files is None and ignore_files is not None:
            # �״��к��Ե�ˢ��
            self.ignored_assigned_files = ignore_files
        elif self.ignored_assigned_files is not None and ignore_files is None:
            resp = QMessageBox.question(self.__parent, '�Ƿ���������嵥', '\n'.join(self.ignored_assigned_files),
                                        QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            if resp == QMessageBox.No:
                self.ignored_assigned_files = None
        else:
            self.ignored_assigned_files = ignore_files
        self.assigned_material_dict.clear()
        self.detail_assigned_dict.clear()
        for i in range(self.assigned_list_widget.count()):
            current_item = self.assigned_list_widget.item(i)
            f_name = current_item.text()
            tf = os.path.basename(f_name)
            dot_index = tf.rindex('.')
            product_bill = tf[:dot_index]
            if self.ignored_assigned_files is not None and f_name in self.ignored_assigned_files:
                continue
            try:
                wb = load_workbook(f_name, data_only=True)
                ws_s = wb.sheetnames
                for s in ws_s:
                    ws = wb[s]
                    r_i = 0
                    for r in ws.iter_rows():
                        if r_i == 0:
                            r_i += 1
                            continue
                        p_id_cell = r[0].value
                        if p_id_cell is None:
                            break
                        s1 = float(r[8].value)
                        s2 = float(r[9].value)
                        s3 = float(r[10].value)
                        if p_id_cell in self.detail_assigned_dict:
                            ll: list = self.detail_assigned_dict[p_id_cell]
                            ll.append((product_bill, s, s1, s2, s3))
                        else:
                            self.detail_assigned_dict[p_id_cell] = [(product_bill, s, s1, s2, s3)]
                        if p_id_cell in self.assigned_material_dict:
                            ss = self.assigned_material_dict[p_id_cell]
                            ss[0] += s1
                            ss[1] += s2
                            ss[2] += s3
                        else:
                            self.assigned_material_dict[p_id_cell] = [s1, s2, s3]
                        r_i += 1
            except Exception as ex:
                self.__parent.show_status_message(f'{f_name} -> ����{ex.__str__()}')
                current_item.setBackground(QColor(255, 0, 0))


class NAssemblyToolWindow(QMainWindow, Ui_MainWindow):
    __pre_header = ['ͼʾ', '�����', '����', '����', '��λ']
    __post_header = ['���ϲ���', '�ɹ��ƻ�']

    # __header = ['ͼʾ', '�����', '����', '����', '��λ',
    #             '�е�ERP���ϱ��', '���ÿ��', '��������', '����ERP���ϱ��', '���ÿ��', '��������',
    #             '�����ֳ����', '��������', '���ϲ���', '�ɹ��ƻ�']

    def __init__(self, parent, database, user, **other):
        super(NAssemblyToolWindow, self).__init__(parent)
        self.__database = database
        self.__user = user
        self.__work_as_admin = False
        self.__dir = other['dir'] if 'dir' in other else None
        self.__product_plan_file = other['product'] if 'product' in other else None
        self.__purchase_plan = other['plan'] if 'plan' in other else None
        # �������ϼ���Ĳ�λ������˳�򣬴������ȼ�
        self.__calculate_position = other['position'] if 'position' in other else ''
        # ��ʾ�ֿ���Ĳֿ�
        self.__ref_position = other['ref_storing'] if 'ref_storing' in other else ''
        self.material_list = {}
        self.__import_cache = []  # �������ݵ���ʱ�洢
        self.__import_sheet = ''  # �������ݵ����ݱ�����
        self.__calculated_item = {}  # ���ڼ����������Ŀ��{����ţ�����}
        self.progressStructTree = ProgressStructPanel(self, database)
        self.selectedProcessList = SelectedProcessPanel(self)
        self.materialTab = MaterialTab(self, self.__database, _dir=self.__dir, plan=self.__purchase_plan)
        self.setup_ui()

    def setup_ui(self):
        super(NAssemblyToolWindow, self).setupUi(self)
        m_h_layout = QHBoxLayout(self.centralwidget)
        v_splitter = QSplitter(Qt.Horizontal, self)
        h_splitter = QSplitter(Qt.Vertical, self)
        h_splitter.addWidget(self.progressStructTree)
        h_splitter.addWidget(self.selectedProcessList)
        v_splitter.addWidget(h_splitter)
        v_splitter.addWidget(self.materialTab)
        v_splitter.setStretchFactor(0, 1)
        v_splitter.setStretchFactor(1, 2)
        m_h_layout.addWidget(v_splitter)
        self.centralwidget.setLayout(m_h_layout)
        self.setWindowTitle('�����������Ϲ���')
        if self.__database.get_database_type() == 'SQLite':
            self.__change_platte()
        self.importSiteStoringAction.setVisible(False)
        self.modifyStockAction.setVisible(False)
        # ��Ӧ����
        self.quitAction.triggered.connect(self.close)
        self.addProgressAction.triggered.connect(self.add_progress)
        self.calculationMaterialAction.triggered.connect(self.do_calculation)
        self.importSiteStoringAction.triggered.connect(lambda: self.__import_storing_data())
        self.importZdStoringAction.triggered.connect(lambda: self.__import_storing_data(_type=1))
        self.importJlStoringAction.triggered.connect(lambda: self.__import_storing_data(_type=2))
        self.importJoStoringAction.triggered.connect(lambda: self.__import_storing_data(_type=3))
        self.generateBomAction.triggered.connect(self.generate_bom)
        self.pickupFromJlAction.triggered.connect(lambda: self.pick_up_items(_type=2))
        self.pickupFromZdAction.triggered.connect(lambda: self.pick_up_items(_type=1))
        self.pickupFromSiteAction.triggered.connect(lambda: self.pick_up_items(_type=3))
        self.calculateThrListAction.triggered.connect(self.__calculate_material_thr_list)
        self.adminRightAction.triggered.connect(self.__handle_admin_right)
        self.modifyStockAction.triggered.connect(self.__modify_part_stock)
        self.importZdFoundationDataAction.triggered.connect(self.__import_zd_foundation_data_handler)
        self.importJoFoundationDataAction.triggered.connect(self.__import_jo_foundation_data_handler)
        self.readProductListAction.triggered.connect(self.open_product_file)

        # TODO �������ϱ��ı�ͷ
        pass

    def open_product_file(self):
        try:
            os.startfile(self.__product_plan_file)
        except Exception as ex:
            QMessageBox.warning(self, '�򿪲�Ʒ�嵥ʱ�쳣', ex.__str__())

    def __import_zd_foundation_data_handler(self):
        """
        �����е����ϻ������ݵķ���
        :return:
        """
        caption = 'ѡ���е����ϻ�����Ϣ�ļ�'
        f, f_type = QFileDialog.getOpenFileName(self, caption, filter='Excel Files (*.xls *.xlsx)')
        if f != '':
            if f[-1] == 'x' or f[-1] == 'X':
                excel_file = ExcelHandler3(f)
            else:
                excel_file = ExcelHandler2(f)
            s_names = excel_file.get_sheets_name()
            dd_s = excel_file.get_datas(s_names[0])[1]
            erp_ids = dd_s['���ϱ���']
            descriptions = dd_s['��������']
            units = dd_s['��λ']
            i = 0
            all_data = []
            for erp_id in erp_ids:
                all_data.append((erp_id[1], descriptions[i][1], units[i][1]))
                i += 1
            result = self.__database.update_zd_erp_foundation_info(all_data)
            QMessageBox.information(self, '���', result)

    def __import_jo_foundation_data_handler(self):
        """
        ������ŷ���ϻ������ݵ���Ӧ����
        :return:
        """
        caption = 'ѡ����ŷ���ϻ�����Ϣ�ļ�'
        f, f_type = QFileDialog.getOpenFileName(self, caption, filter='Excel Files (*.xls *.xlsx)')
        if f != '':
            if f[-1] == 'x' or f[-1] == 'X':
                excel_file = ExcelHandler3(f)
            else:
                excel_file = ExcelHandler2(f)
            s_names = excel_file.get_sheets_name()
            dd_s = excel_file.get_datas(s_names[0])[1]
            erp_ids = dd_s['���ϱ���']
            descriptions = dd_s['��������']
            units = dd_s['��λ']
            i = 0
            all_data = []
            for erp_id in erp_ids:
                all_data.append((erp_id[1], descriptions[i][1], units[i][1]))
                i += 1
            result = self.__database.update_jo_erp_foundation_info(all_data)
            QMessageBox.information(self, '���', result)

    def __modify_part_stock(self):
        """
        �޸ĵ������ϵĿ��
        :return:
        """
        all_positions = self.__database.get_all_storing_position()
        pos, ok = QInputDialog.getItem(self, 'ѡ���λ', '��λ��ţ�', all_positions, 0, False)
        if not ok:
            return
        part_id, ok = QInputDialog.getInt(self, '����', '����ţ�', min=1)
        if not ok:
            return
        the_parts = Part.get_parts(self.__database, part_id=part_id)
        if len(the_parts) < 1:
            self.show_status_message('û�и�����ŵ���Ϣ��', 5)
            return
        qty, ok = QInputDialog.getDouble(self, '����', '������', min=0., decimals=2)
        if not ok:
            return
        unit_price, ok = QInputDialog.getDouble(self, '����', 'ȥ˰���ۣ�', value=0., min=0., decimals=2)
        if not ok:
            return
        resp = QMessageBox.question(self, 'ȷ��',
                                    '�ܽ᣺\n����ţ�{0:08d}\n������{1:.2f}\n��λ��{2}\n���ۣ�{3:.2f}'.format(part_id, qty,
                                                                                                         pos,
                                                                                                         unit_price),
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if resp == QMessageBox.Yes:
            today = datetime.datetime.today()
            t_str = today.strftime('%Y-%m-%d')
            self.__database.update_part_storing(part_id, qty, pos, t_str, unit_price)
            self.show_status_message('��ɸ��¡�', 3)

    def __handle_admin_right(self):
        """
        ���ù���ԱȨ�޵�
        :return:
        """
        admin_actions = (self.importSiteStoringAction, self.modifyStockAction)
        do_c = False
        if self.__work_as_admin:
            self.__work_as_admin = False
            do_c = True
            self.adminRightAction.setText('��ȡ����ԱȨ��')
            self.show_status_message('�����ڲ��ǹ���Ա�ˡ�', 5)
        else:
            if self.__database.get_database_type() == 'SQLite':
                self.__work_as_admin = True
                do_c = True
                self.adminRightAction.setText('�رչ���ԱȨ��')
                self.show_status_message('�������ǹ���Ա�ˡ�', 5)
            else:
                text, ok = QInputDialog.getText(self, '����', '����Ա���룺', QLineEdit.Password)
                if ok:
                    if text == 'zd':
                        # ����һЩ�µĲ˵�
                        self.__work_as_admin = True
                        do_c = True
                        self.adminRightAction.setText('�رչ���ԱȨ��')
                        self.show_status_message('������ȷ���������ǹ���Ա�ˡ�', 5)
                    else:
                        self.show_status_message('�������', 5)
        if do_c:
            for a in admin_actions:
                a.setVisible(self.__work_as_admin)

    def __calculate_material_thr_list(self):
        """
        ͨ��ѡ��Excel�嵥���������ϼ���
        :return:
        """
        caption = 'ѡ�������嵥'
        f, f_type = QFileDialog.getOpenFileName(self, caption, filter='Excel Files (*.xls *.xlsx)')
        if f == '':
            return
        dlg = NImportSettingDialog(self, '�趨�����������', self.__database)
        if f[-1] == 'x' or f[-1] == 'X':
            excel_file = ExcelHandler3(f)
        else:
            excel_file = ExcelHandler2(f)
        rows = ('�����', '����')
        dlg.set_excel_mode(rows, excel_file, general_use=True)
        dlg.exec_()
        if len(self.__import_cache) < 1:
            self.show_status_message('û�����ݡ�', 3)
            return
        part_dict = {}
        material_dict = {}
        for r in self.__import_cache:
            if type(r[0]) == str and len(r[0]) < 1:
                continue
            qty = r[1]
            if r[0] in part_dict:
                p = part_dict[r[0]]
                material_dict[p] += qty
            else:
                ps = Part.get_parts(self.__database, part_id=r[0])
                if ps is not None and len(ps) > 0:
                    p = ps[0]
                    part_dict[r[0]] = p
                    material_dict[p] = qty
                else:
                    continue
        m_tab = self.materialTab.fill_data(material_dict)
        self.materialTab.set_tab_name(self.__import_sheet, table_index=m_tab)
        self.__import_sheet = None
        c = len(material_dict)
        QMessageBox.information(self, '', f'һ�������� {c} �����ϡ�')

    def show_status_message(self, msg, sec=None):
        if sec is None:
            self.statusbar.showMessage(msg)
        else:
            self.statusbar.showMessage(msg, sec * 1000)

    def closeEvent(self, event):
        if self.__database is not None:
            self.__database.close()

    def pick_up_items(self, _type):
        """
        �������ϵ�
        :param _type: 1=�е£�2=���֣�3=�ֳ�
        :return:
        """
        try:
            if self.materialTab.current_material_4_pick_up is not None:
                t_file = self.materialTab.current_material_4_pick_up
                if is_used(t_file):
                    raise Exception(f'{t_file}��ռ�ã��޷�������')
            # �������϶Ի���
            material_table, table_txt = self.materialTab.get_current_material_table()
            all_materials = material_table.get_all_material()
            if _type == 1 and len(all_materials[2]) < 1:
                QMessageBox.warning(self, '�жϲ���', 'û�д��е²ֿ����õ����ϡ�')
                return
            if _type == 2 and len(all_materials[3]) < 1:
                QMessageBox.warning(self, '�жϲ���', 'û�дӾ��ֲֿ����õ����ϡ�')
                return
            if _type == 3 and len(all_materials[1]) < 1:
                QMessageBox.warning(self, '�жϲ���', 'û�д������ֳ����õ����ϡ�')
                return
            pick_material_dialog = NCreatePickBillDialog(self.__database, parent=self)
            the_config_dict = {}
            the_text, ok = QInputDialog.getText(self, 'Ĭ�Ϻ�ͬ��', '��ͬ�ţ�')
            if ok:
                try_record = self.__database.get_products_by_id(the_text)
                if len(try_record) > 0:
                    the_config_dict['��ͬ��'] = the_text
                else:
                    QMessageBox.warning(self, '��Ч����', '������Ĳ�Ʒ��Ų����ڣ�')
            else:
                self.show_status_message('û��ѡ���ͬ��', 5)
            if _type == 3:
                the_config_dict['�ֿ�'] = 'A'
                handle_material = all_materials[1]
            else:
                the_config_dict['�ֿ�'] = 'D' if _type == 1 else 'F'
                handle_material = all_materials[_type + 1]
            if self.__user is not None and len(self.__user) > 0:
                the_config_dict['������'] = self.__user
            today_bill = QDate.currentDate().toString('yyMMdd')
            the_config_dict['�嵥'] = self.__database.get_available_supply_operation_bill(prefix=f'P{today_bill}')
            pick_material_dialog.set_config(the_config_dict)
            items = []
            part_dict = {}  # ���ڵǼ��������ݵ�ӳ���
            for r in handle_material:
                p: Part = r[0]
                erp_id = r[1]
                qty = r[2]
                info_from_erp = False
                des_str = None
                unit_str = None
                if _type != 3:
                    info = self.__database.get_erp_info(erp_id, jl_erp=_type == 2)
                    if info is not None:
                        des_str = info[1]
                        unit_str = info[2]
                        info_from_erp = True
                if not info_from_erp:
                    description = [p.name]
                    if p.description is not None and p.description != '':
                        description.append(p.description)
                    t1 = p.get_specified_tag(self.__database, 'Ʒ��')
                    t2 = p.get_specified_tag(self.__database, '��׼')
                    t3 = p.get_specified_tag(self.__database, '��λ')
                    unit_str = t3 if len(t3) > 0 else '��'
                    if len(t1) > 0:
                        description.append(f'({t1})')
                    if len(t2) > 0:
                        description.append(f'({t2})')
                    des_str = ' '.join(description)
                the_part_id = p.get_part_id()
                items.append([p.part_id, erp_id, des_str, unit_str, qty])
                part_dict[the_part_id] = qty
            pick_material_dialog.add_items(items)
            resp = pick_material_dialog.exec_()
            if resp == 0 and self.materialTab.current_material_4_pick_up is not None:
                # ��¼�������ݡ��޸�Excel�ļ����Ǽǳ����¼��
                pick_column = 13 + _type
                assigned_colum = 7 + _type
                wb = load_workbook(self.materialTab.current_material_4_pick_up)
                ws = wb[table_txt]
                r_i = 0
                for r in ws.iter_rows():
                    if r_i == 0:
                        r_i += 1
                        continue
                    p_id_v = r[0].value
                    if p_id_v is None:
                        break
                    if p_id_v in part_dict:
                        o_data = float(r[pick_column].value)
                        a_data = float(r[assigned_colum].value)
                        r[pick_column].value = o_data + part_dict[p_id_v]
                        r[assigned_colum].value = a_data - part_dict[p_id_v]
                wb.save(self.materialTab.current_material_4_pick_up)
                self.show_status_message(f'������{self.materialTab.current_material_4_pick_up}�ļ���{table_txt}���ݱ�')
        except Exception as ex:
            QMessageBox.warning(self, '', f'����ʱ�쳣��{ex.__str__()}')

    def generate_bom(self):
        try:
            materialTable, table_txt = self.materialTab.get_current_material_table()
            all_materials = materialTable.get_all_material()
            bom_type_list = ('ͳ���嵥', 'װ���¼�嵥', '�����嵥')
            bom_type, ok = QInputDialog.getItem(self, 'ѡ��', '�嵥��ʽ', bom_type_list, current=0, editable=False)
            if not ok:
                self.show_status_message('ȡ������嵥��', 5)
                return
            not_enough_qty_item = all_materials[0]
            if not_enough_qty_item > 0:
                resp = QMessageBox.question(self, 'ȱ��', f'�� {not_enough_qty_item} �����ϵĿ��ÿ�治�㣬�Ƿ������',
                                            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if resp == QMessageBox.No:
                    self.show_status_message('ȡ������嵥��', 5)
                    return

            wb = xw.Book()
            des_dict = {}

            bom_type_index = bom_type_list.index(bom_type)

            if bom_type_index == 0:
                if len(all_materials[4]) < 1:
                    QMessageBox.warning(self, '', 'û�������嵥', QMessageBox.Yes)
                    return
                zd_storage_dict = dict()
                for m in all_materials[2]:
                    zd_storage_dict[m[0]] = m[2]
                jl_storage_dict = dict()
                for m in all_materials[3]:
                    jl_storage_dict[m[0]] = m[2]
                site_storage_dict = dict()
                for m in all_materials[1]:
                    site_storage_dict[m[0]] = m[2]
                # ����ͳ���嵥
                ws = wb.sheets.active
                ws.name = 'bom' if table_txt is None else table_txt
                headers = (
                    '�����', '����', '�ͺ�����', 'Ʒ��', '��׼', '��λ', '����', '�����е����ϱ���', '�е²ֿ���',
                    '���ֲֿ���', '�ֳ����', '����', '���ϲ���', '��治��', 'P1', 'P2', 'P3', 'R')
                ws.range((1, 1), (1, len(headers) + 1)).value = headers
                i = 2
                for r in all_materials[4]:
                    p: Part = r[0]
                    part_id = p.get_part_id()
                    id_cell = ws.range((i, 1))
                    id_cell.value = part_id
                    id_cell.number_format = '00000000'
                    qty_zd = zd_storage_dict[p] if p in zd_storage_dict.keys() else 0.0
                    qty_jl = jl_storage_dict[p] if p in jl_storage_dict.keys() else 0.0
                    qty_site = site_storage_dict[p] if p in site_storage_dict.keys() else 0.0
                    name_cell = ws.range((i, 2))
                    name_cell.value = p.name
                    description_cell = ws.range((i, 3))
                    description_cell.value = p.description
                    t1 = p.get_specified_tag(self.__database, 'Ʒ��')
                    brand_cell = ws.range((i, 4))
                    brand_cell.value = t1
                    t2 = p.get_specified_tag(self.__database, '��׼')
                    standard_cell = ws.range((i, 5))
                    standard_cell.value = t2
                    t3 = p.get_specified_tag(self.__database, '��λ')
                    unit_cell = ws.range((i, 6))
                    unit_cell.value = t3
                    qty_cell = ws.range((i, 7))
                    qty_cell.value = r[1]
                    qty_cell.number_format = '0.00'
                    t4 = p.get_specified_tag(self.__database, '�����е�ERP���ϱ���')
                    zd_erp_id_cell = ws.range((i, 8))
                    zd_erp_id_cell.value = t4
                    zd_qty_cell = ws.range((i, 9))
                    zd_qty_cell.value = qty_zd
                    zd_qty_cell.number_format = '0.00'
                    jl_qty_cell = ws.range((i, 10))
                    jl_qty_cell.value = qty_jl
                    jl_qty_cell.number_format = '0.00'
                    site_qty_cell = ws.range((i, 11))
                    site_qty_cell.value = qty_site
                    site_qty_cell.number_format = '0.00'
                    t5 = p.get_specified_tag(self.__database, '���')
                    type_cell = ws.range((i, 12))
                    type_cell.value = t5
                    t6 = p.get_specified_tag(self.__database, '���ϲ���')
                    supply_type_cell = ws.range((i, 13))
                    supply_type_cell.value = t6
                    warning_cell = ws.range((i, 14))
                    warning_cell.formula = '=SUM(I{0}:K{0})-R{0}'.format(i)
                    warning_cell.number_format = '0.00'
                    p_cell = [ws.range((i, 15)), ws.range((i, 16)), ws.range((i, 17))]
                    for c in p_cell:
                        c.value = 0.0
                        c.number_format = '0.00'
                    r_cell = ws.range((i, 18))
                    r_cell.formula = '=G{0}-SUM(O{0}:Q{0})'.format(i)
                    r_cell.number_format = '0.00'
                    i += 1
            else:
                if bom_type_index == 2:
                    self.show_status_message('�����ɶ���嵥��', 10.0)

                    # �����ֳ�������
                    ws1 = wb.sheets.active
                    ws1.name = '�ֳ�����'
                    headers = ('���', '�����', 'ͼʾ', '����', '����', '��λ')
                    ws1.range((1, 1), (1, len(headers) + 1)).value = headers

                    c_m_1 = len(all_materials[1])
                    c_row = 1 if c_m_1 < 1 else c_m_1 + 1
                    ws1.range((1, 1), (c_row, 1)).column_width = 5
                    ws1.range((1, 2), (c_row, 2)).column_width = 10
                    ws1.range((1, 3), (c_row, 3)).column_width = 20
                    ws1.range((1, 4), (c_row, 4)).column_width = 35
                    ws1.range((1, 5), (c_row, 5)).column_width = 8
                    ws1.range((1, 6), (c_row, 6)).column_width = 5
                    if c_m_1 > 0:
                        ws1.range((2, 1), (c_row, 1)).row_height = 80
                        i = 2
                        img_name = {}
                        for r in all_materials[1]:
                            p: Part = r[0]
                            part_id = p.get_part_id()
                            ws1.range((i, 1)).value = i - 1
                            id_cell = ws1.range((i, 2))
                            id_cell.value = part_id
                            id_cell.number_format = '00000000'
                            img = self.__database.generate_a_image(p.get_part_id())
                            if img is not None:
                                img_w, img_h = Image.open(img).size
                                img_cell = ws1.range((i, 3))
                                cell_ratio = img_cell.width / img_cell.height
                                img_ratio = img_w / img_h
                                if img_ratio > cell_ratio:
                                    pic_width = img_cell.width - 2
                                    pic_height = pic_width / img_ratio
                                else:
                                    pic_height = img_cell.height - 2
                                    pic_width = pic_height * img_ratio
                                f = pic_height / img_h
                                pic_top = img_cell.top + (img_cell.height - pic_height) / 2
                                pic_left = img_cell.left + (img_cell.width - pic_width) / 2
                                if p.part_id in img_name:
                                    _name = f'{0}_{img_name[p.part_id] + 1}'
                                    img_name[p.part_id] += 1
                                else:
                                    _name = p.part_id
                                    img_name[_name] = 1
                                ws1.pictures.add(img, left=pic_left, top=pic_top, width=pic_width, height=pic_height,
                                                 scale=int(f), name=_name)
                            if part_id not in des_dict:
                                description = [p.name]
                                if p.description is not None and p.description != '':
                                    description.append(p.description)
                                t1 = p.get_specified_tag(self.__database, 'Ʒ��')
                                t2 = p.get_specified_tag(self.__database, '��׼')
                                t3 = p.get_specified_tag(self.__database, '��λ')
                                unit_str = t3 if len(t3) > 0 else '��'
                                if len(t1) > 0:
                                    description.append(f'({t1})')
                                if len(t2) > 0:
                                    description.append(f'({t2})')
                                des_str = ' '.join(description)
                                des_dict[part_id] = (des_str, unit_str)
                            else:
                                des_str, unit_str = des_dict[part_id]
                            ws1.range((i, 4)).value = des_str
                            ws1.range((i, 5)).value = r[2]
                            ws1.range((i, 6)).value = unit_str
                            i += 1

                    # �����е²ֿ������
                    ws2 = wb.sheets.add('�е²ֿ�����', after=ws1)
                    headers = ('���', '�����', 'ERP���ϱ��', '��������', '����', '��λ')
                    ws2.range((1, 1), (1, len(headers) + 1)).value = headers
                    c_m_2 = len(all_materials[2])
                    c_row = 1 if c_m_2 > 0 else c_m_2 + 1
                    ws2.range((1, 1), (c_row, 1)).column_width = 5
                    ws2.range((1, 2), (c_row, 2)).column_width = 10
                    ws2.range((1, 3), (c_row, 3)).column_width = 13
                    ws2.range((1, 4), (c_row, 4)).column_width = 60
                    ws2.range((1, 5), (c_row, 5)).column_width = 5
                    ws2.range((1, 6), (c_row, 6)).column_width = 5
                    if c_m_2 > 0:
                        i = 2
                        for r in all_materials[2]:
                            p: Part = r[0]
                            part_id = p.get_part_id()
                            erp_id = r[1]
                            ws2.range((i, 1)).value = i - 1
                            id_cell = ws2.range((i, 2))
                            id_cell.value = part_id
                            id_cell.number_format = '00000000'
                            ws2.range((i, 3)).value = erp_id
                            info = self.__database.get_erp_info(erp_id)
                            if info is not None:
                                ws2.range((i, 4)).value = info[1]
                                ws2.range((i, 6)).value = info[2]
                            else:
                                if part_id in des_dict:
                                    p_des, unit_str = des_dict[part_id]
                                    ws2.range((i, 4)).value = p_des
                                    ws2.range((i, 6)).value = unit_str
                                else:
                                    description = [p.name]
                                    if p.description is not None and p.description != '':
                                        description.append(p.description)
                                    t1 = p.get_specified_tag(self.__database, 'Ʒ��')
                                    t2 = p.get_specified_tag(self.__database, '��׼')
                                    t3 = p.get_specified_tag(self.__database, '��λ')
                                    unit_str = t3 if len(t3) > 0 else '��'
                                    if len(t1) > 0:
                                        description.append(f'({t1})')
                                    if len(t2) > 0:
                                        description.append(f'({t2})')
                                    des_str = ' '.join(description)
                                    des_dict[part_id] = (des_str, unit_str)
                                    ws2.range((i, 4)).value = des_str
                                    ws2.range((i, 6)).value = unit_str
                            ws2.range((i, 5)).value = r[2]
                            i += 1

                    # ������ֲֿ������
                    ws3 = wb.sheets.add('���ֲֿ�����', after=ws2)
                    ws3.range((1, 1), (1, len(headers) + 1)).value = headers
                    c_m_3 = len(all_materials[3])
                    c_row = 1 if c_m_3 > 0 else c_m_3 + 1
                    ws3.range((1, 1), (c_row, 1)).column_width = 5
                    ws3.range((1, 2), (c_row, 2)).column_width = 10
                    ws3.range((1, 3), (c_row, 3)).column_width = 13
                    ws3.range((1, 4), (c_row, 4)).column_width = 60
                    ws3.range((1, 5), (c_row, 5)).column_width = 5
                    ws3.range((1, 6), (c_row, 6)).column_width = 5
                    if c_m_3 > 0:
                        i = 2
                        for r in all_materials[3]:
                            p: Part = r[0]
                            part_id = p.get_part_id()
                            erp_id = r[1]
                            ws3.range((i, 1)).value = i - 1
                            id_cell = ws3.range((i, 2))
                            id_cell.value = part_id
                            id_cell.number_format = '00000000'
                            ws3.range((i, 3)).value = erp_id
                            info = self.__database.get_erp_info(erp_id, jl_erp=True)
                            if info is not None:
                                ws3.range((i, 4)).value = info[1]
                                ws3.range((i, 6)).value = info[2]
                            else:
                                if part_id in des_dict:
                                    p_des, unit_str = des_dict[part_id]
                                    ws3.range((i, 4)).value = p_des
                                    ws3.range((i, 6)).value = unit_str
                                else:
                                    description = [p.name]
                                    if p.description is not None and p.description != '':
                                        description.append(p.description)
                                    t1 = p.get_specified_tag(self.__database, 'Ʒ��')
                                    t2 = p.get_specified_tag(self.__database, '��׼')
                                    t3 = p.get_specified_tag(self.__database, '��λ')
                                    unit_str = t3 if len(t3) > 0 else '��'
                                    if len(t1) > 0:
                                        description.append(f'({t1})')
                                    if len(t2) > 0:
                                        description.append(f'({t2})')
                                    des_str = ' '.join(description)
                                    des_dict[part_id] = (des_str, unit_str)
                                    ws3.range((i, 4)).value = des_str
                                    ws3.range((i, 6)).value = unit_str
                            ws3.range((i, 5)).value = r[2]
                            i += 1

                    # �����İ��Ʒ
                    ws4 = wb.sheets.add('���Ʒ', after=ws3)
                    output_items = self.selectedProcessList.get_all_item()
                    headers = ('���', '�����', 'ͼʾ', '����', '����', '��λ')
                    ws4.range((1, 1), (1, len(headers) + 1)).value = headers
                    c_m_4 = len(output_items)
                    c_row = 1 if c_m_4 < 1 else c_m_4 + 1
                    ws4.range((1, 1), (c_row, 1)).column_width = 5
                    ws4.range((1, 2), (c_row, 2)).column_width = 10
                    ws4.range((1, 3), (c_row, 3)).column_width = 20
                    ws4.range((1, 4), (c_row, 4)).column_width = 35
                    ws4.range((1, 5), (c_row, 5)).column_width = 8
                    ws4.range((1, 6), (c_row, 6)).column_width = 5
                    if c_m_4 > 0:
                        ws4.range((2, 1), (c_row, 1)).row_height = 80
                        i = 2
                        for r in output_items:
                            p: Part = r[0]
                            part_id = p.get_part_id()
                            ws4.range((i, 1)).value = i - 1
                            id_cell = ws4.range((i, 2))
                            id_cell.value = part_id
                            id_cell.number_format = '00000000'
                            img = self.__database.generate_a_image(p.get_part_id())
                            if img is not None:
                                img_w, img_h = Image.open(img).size
                                img_cell = ws4.range((i, 3))
                                cell_ratio = img_cell.width / img_cell.height
                                img_ratio = img_w / img_h
                                if img_ratio > cell_ratio:
                                    pic_width = img_cell.width - 2
                                    pic_height = pic_width / img_ratio
                                else:
                                    pic_height = img_cell.height - 2
                                    pic_width = pic_height * img_ratio
                                f = pic_height / img_h
                                pic_top = img_cell.top + (img_cell.height - pic_height) / 2
                                pic_left = img_cell.left + (img_cell.width - pic_width) / 2
                                ws4.pictures.add(img, left=pic_left, top=pic_top, width=pic_width, height=pic_height,
                                                 scale=int(f), name=p.part_id)
                            if part_id not in des_dict:
                                description = [p.name]
                                if p.description is not None and p.description != '':
                                    description.append(p.description)
                                t1 = p.get_specified_tag(self.__database, 'Ʒ��')
                                t2 = p.get_specified_tag(self.__database, '��׼')
                                t3 = p.get_specified_tag(self.__database, '��λ')
                                unit_str = t3 if len(t3) > 0 else '��'
                                if len(t1) > 0:
                                    description.append(f'({t1})')
                                if len(t2) > 0:
                                    description.append(f'({t2})')
                                des_str = ' '.join(description)
                                des_dict[part_id] = (des_str, unit_str)
                            else:
                                des_str, unit_str = des_dict[part_id]
                            ws4.range((i, 4)).value = des_str
                            ws4.range((i, 5)).value = r[1]
                            ws4.range((i, 6)).value = unit_str
                            i += 1

                    # ��������
                    ws5 = wb.sheets.add('ȫ������', after=ws4)
                else:
                    ws5 = wb.sheets.active
                    ws5.name = 'ȫ������'

                if ws5 is None:
                    raise Exception('���ɡ�ȫ�����ϡ��Ĺ������쳣��')

                _txt, _ok = QInputDialog.getText(self, '����', '��ȫ�����ϡ����������')

                headers = ('���', '�����', 'ͼʾ', '��������', '����', '���ÿ��', '��λ', '���ϲ���')
                header_cells = ws5.range((1, 1), (1, len(headers)))
                header_cells.value = headers
                header_cells.api.HorizontalAlignment = -4108
                header_cells.api.Borders(9).LineStyle = 1  # �ײ��߿�
                header_cells.api.Borders(9).Weight = 3
                header_cells.api.Font.Bold = True  # ����Ӵ�
                c_m_5 = len(all_materials[4])
                c_row = 1 if c_m_5 < 1 else c_m_5 + 1
                ws5.range((1, 1), (c_row, 1)).column_width = 5
                ws5.range((1, 2), (c_row, 2)).column_width = 10
                ws5.range((1, 3), (c_row, 3)).column_width = 20
                ws5.range((1, 4), (c_row, 4)).column_width = 40
                ws5.range((1, 5), (c_row, 5)).column_width = 10
                ws5.range((1, 6), (c_row, 6)).column_width = 10
                ws5.range((1, 7), (c_row, 7)).column_width = 5
                ws5.range((1, 8), (c_row, 8)).column_width = 10
                if c_m_5 > 0:
                    ws5.range((1, 1)).row_height = 20
                    ws5.range((2, 1), (c_row, 1)).row_height = 80
                    i = 2
                    img_name = {}
                    for r in all_materials[4]:
                        p: Part = r[0]
                        part_id = p.get_part_id()
                        qty = r[1]
                        available_qty = r[2]
                        index_cell = ws5.range((i, 1))
                        index_cell.value = i - 1
                        index_cell.api.HorizontalAlignment = -4108
                        id_cell = ws5.range((i, 2))
                        id_cell.value = part_id
                        id_cell.number_format = '00000000'
                        id_cell.api.HorizontalAlignment = -4108
                        img = self.__database.generate_a_image(p.get_part_id())
                        if img is not None:
                            img_w, img_h = Image.open(img).size
                            img_cell = ws5.range((i, 3))
                            cell_ratio = img_cell.width / img_cell.height
                            img_ratio = img_w / img_h
                            if img_ratio > cell_ratio:
                                pic_width = img_cell.width - 2
                                pic_height = pic_width / img_ratio
                            else:
                                pic_height = img_cell.height - 2
                                pic_width = pic_height * img_ratio
                            f = pic_height / img_h
                            pic_top = img_cell.top + (img_cell.height - pic_height) / 2
                            pic_left = img_cell.left + (img_cell.width - pic_width) / 2
                            if p.part_id in img_name:
                                _name = f'{0}_{img_name[p.part_id] + 1}'
                                img_name[p.part_id] += 1
                            else:
                                _name = p.part_id
                                img_name[_name] = 1
                            ws5.pictures.add(img, left=pic_left, top=pic_top, width=pic_width, height=pic_height,
                                             scale=int(f), name=_name)
                        qty_cell = ws5.range((i, 5))
                        qty_cell.value = qty
                        qty_cell.number_format = '0.00'
                        qty_cell.api.HorizontalAlignment = -4108
                        available_qty_cell = ws5.range((i, 6))
                        available_qty_cell.value = available_qty
                        available_qty_cell.number_format = '0.00'
                        available_qty_cell.api.HorizontalAlignment = -4108
                        if part_id in des_dict:
                            des_str, unit_str = des_dict[part_id]
                        else:
                            description = [p.name]
                            if p.description is not None and p.description != '':
                                description.append(p.description)
                            t1 = p.get_specified_tag(self.__database, 'Ʒ��')
                            t2 = p.get_specified_tag(self.__database, '��׼')
                            t3 = p.get_specified_tag(self.__database, '��λ')
                            unit_str = t3 if len(t3) > 0 else '��'
                            if len(t1) > 0:
                                description.append(f'({t1})')
                            if len(t2) > 0:
                                description.append(f'({t2})')
                            des_str = '\n'.join(description)
                            des_dict[part_id] = (des_str, unit_str)
                        des_cell = ws5.range((i, 4))
                        des_cell.value = des_str
                        des_cell.api.HorizontalAlignment = -4108
                        unit_cell = ws5.range((i, 7))
                        unit_cell.value = unit_str
                        unit_cell.api.HorizontalAlignment = -4108
                        t4 = p.get_specified_tag(self.__database, '���ϲ���')
                        if len(t4) > 0:
                            supply_type_cell = ws5.range((i, 8))
                            supply_type_cell.value = t4
                            supply_type_cell.api.HorizontalAlignment = -4108
                        i += 1
                    # ҳ������
                    ws5.page_setup.print_area = f'$A$1:$H${c_row}'
                    ws5.api.PageSetup.PrintTitleRows = '$1:$1'
                    ws5.api.PageSetup.Zoom = 72
                    ws5.api.PageSetup.CenterHorizontally = True
                    if _ok:
                        ws5.api.PageSetup.CenterHeader = _txt
                    _today = datetime.datetime.today()
                    ws5.api.PageSetup.RightHeader = _today.strftime('%Y/%m/%d %H:%M:%S')
                    ws5.api.PageSetup.CenterFooter = r'�� &P ҳ���� &N ҳ'
            QMessageBox.information(self, '', '����嵥�����ɡ�')
        except Exception as ex:
            QMessageBox.warning(self, '�쳣', ex.__str__())

    def add_progress(self):
        selected_items = self.progressStructTree.get_selected_item()
        if selected_items is None:
            self.statusbar.showMessage('û��ѡ����')
            return
        qty, ok = QInputDialog.getInt(self, '����', '����', value=1)
        if ok:
            item_list = []
            for i in selected_items:
                p = i[0]
                n_qty = i[1] * qty
                item_list.append((p, n_qty))
            self.selectedProcessList.add_row(item_list)
            self.statusbar.showMessage('������ {0} ����Ŀ��'.format(len(item_list)))

    def do_calculation(self):
        """
        ��������
        :return:
        """
        selected_item = {}
        item_list = self.selectedProcessList.get_all_item()
        if len(item_list) < 1:
            self.statusbar.showMessage('û��ѡ����Ŀ��')
            return
        self.show_status_message('���������С���')
        # ���ظ�����Ŀ�ϲ�
        for i in item_list:
            p = i[0]
            q = i[1]
            if p not in selected_item:
                selected_item[p] = q
            else:
                selected_item[p] += q
        self.material_list.clear()
        # Ԥ��Ҫ�Ƴ��ġ�������������Ŀ
        for k in selected_item.keys():
            p_id = k.get_part_id()
            self.__calculated_item[p_id] = selected_item[k]
        for d in selected_item:
            self.__do_statistics(d, selected_item[d])
        # �Ƴ�KBN
        if self.noKbnAction.isChecked():
            neu_dict = {}
            for p in self.material_list.keys():
                t = p.get_specified_tag(self.__database, '���ϲ���')
                if t == 'KBN':
                    continue
                else:
                    neu_dict[p] = self.material_list[p]
            self.material_list = neu_dict
        self.materialTab.fill_data(self.material_list)
        QMessageBox.information(self, '', f'������ɣ����� {len(self.material_list)} �')

    def __do_statistics(self, p: Part, qty):
        children = p.get_children(self.__database)
        if children is None:
            self.__add_2_result(p, qty)
            return
        for c in children:
            p = c[1]
            p_id = p.get_part_id()
            c_qty = qty * c[3]
            # �Բ�ͬ�׶εĽ���(�����ι���)��Ŀ��������
            if p_id in self.__calculated_item:
                r_qty = self.__calculated_item[p_id]
                if c_qty <= r_qty:
                    r_qty -= c_qty
                    if r_qty <= 0.:
                        self.__calculated_item.pop(p_id)
                    else:
                        self.__calculated_item[p_id] = r_qty
                    continue
                else:
                    c_qty -= r_qty
            pur_type = p.get_specified_tag(self.__database, '��Դ')
            # �����������ý����ж�
            p_type = p.get_specified_tag(self.__database, '���')
            # ͳ��Ҫװ���
            if p_type == 'ͼֽ' or p_type == '�ĵ�' or p_type == '���ⵥԪ':
                continue
            # ע��װ�䡢���ƻ�ɹ�ʱ���Ͳ�Ҫ�����²�ѯ
            if pur_type == 'װ��' or pur_type == '����' or pur_type == '�ɹ�':
                self.__add_2_result(p, c_qty)
                continue
            self.__do_statistics(p, c_qty)

    def __add_2_result(self, part, qty):
        if qty <= 0.:
            return
        if part in self.material_list:
            self.material_list[part] += qty
        else:
            self.material_list[part] = qty

    def fill_import_cache(self, data_s, sheet_name=None):
        self.__import_cache.clear()
        self.__import_cache.extend(data_s)
        self.__import_sheet = sheet_name

    def __import_storing_data(self, _type=0):
        """
        ����Excel��ʽ�Ĳִ�����
        :param _type: 0=�ֳ���1=�е²ֿ⣬2=���ֲֿ⣬3=��ŷ�ֿ�
        :return:
        """
        if _type == 0:
            QMessageBox.warning(self, '����', 'Ҫ�ı��ֳ����������ݣ������������')
            caption = 'ѡ���ֳ���������ļ�'
            f, f_type = QFileDialog.getOpenFileName(self, caption, filter='Excel Files (*.xls *.xlsx)')
            if f == '':
                return
            resp = QMessageBox.question(self, '���ݷ�ʽ', '��ǰ����������������',
                                        QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel, QMessageBox.Cancel)
            if resp == QMessageBox.Cancel:
                self.show_status_message('���������ݣ�', 3)
                return
            # if resp == QMessageBox.No:
            #     resp2 = QMessageBox.question( self, 'ȷ��', '�Ƿ�Ҫ���֮ǰ��¼���ֳ���������ݣ�',
            #                                   QMessageBox.Yes | QMessageBox.No, QMessageBox.No )
            #     if resp2 == QMessageBox.Yes:
            #         self.__database.clear_storing_position( 'A' )
            dlg = NImportSettingDialog(self, '�����ֳ��ִ�����', self.__database)
            if f[-1] == 'x' or f[-1] == 'X':
                excel_file = ExcelHandler3(f)
            else:
                excel_file = ExcelHandler2(f)
            rows = ('�����', '����', 'ȥ˰����')
            dlg.set_excel_mode(rows, excel_file, general_use=True)
            dlg.exec_()
            if len(self.__import_cache) < 0:
                self.show_status_message('û�����ݡ�', 3)
                return
            _today = datetime.datetime.today()
            last_picked_date = _today.strftime('%Y-%m-%d')
            c = 0
            for d in self.__import_cache:
                p_id = d[0]  # �����
                qty = d[1]  # Ҫ�ı������
                pp = d[2]  # ȥ˰����
                if p_id is None or (type(p_id) == str and len(p_id) < 1):
                    continue
                if qty is None or (type(qty) == str and len(qty) < 1):
                    continue
                if type(p_id) == str:
                    p_id = int(float(p_id))
                if type(qty) == str:
                    qty = float(qty)
                if pp is None or (type(pp) == str and len(pp) < 1):
                    pp = 0.
                else:
                    if type(pp) == str:
                        pp = float(pp)
                if resp == QMessageBox.Yes:
                    ps = self.__database.get_storing(p_id, 'A')
                    if ps is None or len(ps) < 1:
                        o_qty = 0.  # ԭ�е�����
                    else:
                        o_qty = ps[0][2]
                        pp = ps[0][4]
                    n_qty = o_qty + qty
                else:
                    n_qty = qty
                self.__database.update_part_storing(p_id, n_qty, 'A', last_picked_date, pp)
                c += 1
            QMessageBox.information(self, '������', '�������� {0} �����ݣ�'.format(c))
        elif _type == 1 or _type == 2 or _type == 3:
            _id_str = ''
            try:
                caption = 'ѡ���е²ִ������ļ�'
                if _type == 2:
                    caption = 'ѡ����֣������˲֣��ִ������ļ�'
                else:
                    caption = 'ѡ����ŷ�ִ������ļ�'
                f, f_type = QFileDialog.getOpenFileName(self, caption, filter='Excel Files (*.xls *.xlsx)')
                if f != '':
                    if f[-1] == 'x' or f[-1] == 'X':
                        excel_file = ExcelHandler3(f)
                    else:
                        excel_file = ExcelHandler2(f)
                    s_names = excel_file.get_sheets_name()
                    dd_s = excel_file.get_datas(s_names[0])[1]
                    if _type == 1:
                        p_tag = self.__database.get_tags(name='�����е�ERP���ϱ���')
                    elif _type == 2:
                        p_tag = self.__database.get_tags(name='��������ERP���ϱ���')
                    else:
                        p_tag = self.__database.get_tags(name='��ŷERP���ϱ���')
                    p_tag_id = p_tag[0][0]
                    erp_ids = dd_s['���Ϻ�']
                    qty_s = dd_s['���ÿ����']
                    storage_name_s = dd_s['�ⷿ']
                    last_picked_date_s = dd_s['����������']
                    unit_price_s = dd_s['����']
                    n = 0
                    # ׼���Բ�ͬλ�õ��ۼ�
                    imported_record = {}
                    if _type == 1:
                        self.__database.clear_storing_position('D')
                        self.__database.clear_storing_position('E')
                    elif _type == 2:
                        self.__database.clear_storing_position('F')
                    else:
                        self.__database.clear_storing_position('J')
                    for _id in erp_ids:
                        r_i = _id[0]
                        _id_str = _id[1]
                        erp_id_tag_id = self.__database.get_tags(name=_id_str, parent_id=p_tag_id)
                        if len(erp_id_tag_id) < 1:
                            continue
                        p_s = Part.get_parts_from_tag(self.__database, erp_id_tag_id[0][0])
                        if len(p_s) < 1:
                            continue
                        p: Part = p_s[0]
                        qty = float(qty_s[r_i - 2][1])
                        storage_name = storage_name_s[r_i - 2][1]
                        if _type == 1:
                            storage_name = 'D' if storage_name == 'ԭ��C��' else 'E'
                        elif _type == 2:
                            storage_name = 'F'
                        else:
                            storage_name = 'J'
                        last_picked_date = last_picked_date_s[r_i - 2][1]
                        if last_picked_date == '':
                            _today = datetime.datetime.today()
                            last_picked_date = _today.strftime('%Y-%m-%d')
                        unit_price_ss = unit_price_s[r_i - 2][1]
                        if len(unit_price_ss) < 1:
                            unit_price_ss = '0.0'
                        unit_price = float(unit_price_ss)
                        if storage_name in imported_record:
                            p_id = p.get_part_id()
                            if p_id in imported_record[storage_name]:
                                n_qty = imported_record[storage_name][p_id]
                                n_qty += qty
                                imported_record[storage_name][p_id] = n_qty
                                print(f'{p_id} do sum in different record.')
                            else:
                                imported_record[storage_name][p_id] = qty
                                n_qty = qty
                        else:
                            imported_record[storage_name] = {}
                            n_qty = qty
                        print(f'{p.get_part_id()} {n_qty} {storage_name} {last_picked_date} {unit_price}')
                        self.__database.update_part_storing(p.get_part_id(), n_qty, storage_name, last_picked_date,
                                                            unit_price)
                        n += 1
                    QMessageBox.information(self, '������', '�������� {0} �����ݣ�'.format(n))
            except Exception as ex:
                QMessageBox.critical(self, '�����ж�', f'�ڴ���{_id_str}ʱ�������޷�����Ĵ���')
                raise ex

    def __change_platte(self):
        """
        �ı�����ʱ�����ڵ�����
        :return:
        """
        palette = QPalette()
        brush = QBrush(QColor(0, 0, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Active, QPalette.WindowText, brush)
        brush = QBrush(QColor(255, 170, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Active, QPalette.Button, brush)
        brush = QBrush(QColor(255, 213, 127))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Active, QPalette.Light, brush)
        brush = QBrush(QColor(255, 191, 63))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Active, QPalette.Midlight, brush)
        brush = QBrush(QColor(127, 85, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Active, QPalette.Dark, brush)
        brush = QBrush(QColor(170, 113, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Active, QPalette.Mid, brush)
        brush = QBrush(QColor(0, 0, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Active, QPalette.Text, brush)
        brush = QBrush(QColor(255, 255, 255))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Active, QPalette.BrightText, brush)
        brush = QBrush(QColor(0, 0, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Active, QPalette.ButtonText, brush)
        brush = QBrush(QColor(255, 255, 255))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Active, QPalette.Base, brush)
        brush = QBrush(QColor(255, 170, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Active, QPalette.Window, brush)
        brush = QBrush(QColor(0, 0, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Active, QPalette.Shadow, brush)
        brush = QBrush(QColor(255, 212, 127))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Active, QPalette.AlternateBase, brush)
        brush = QBrush(QColor(255, 255, 220))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Active, QPalette.ToolTipBase, brush)
        brush = QBrush(QColor(0, 0, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Active, QPalette.ToolTipText, brush)
        brush = QBrush(QColor(0, 0, 0, 128))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Active, QPalette.PlaceholderText, brush)
        brush = QBrush(QColor(0, 0, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Inactive, QPalette.WindowText, brush)
        brush = QBrush(QColor(255, 170, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Inactive, QPalette.Button, brush)
        brush = QBrush(QColor(255, 213, 127))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Inactive, QPalette.Light, brush)
        brush = QBrush(QColor(255, 191, 63))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Inactive, QPalette.Midlight, brush)
        brush = QBrush(QColor(127, 85, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Inactive, QPalette.Dark, brush)
        brush = QBrush(QColor(170, 113, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Inactive, QPalette.Mid, brush)
        brush = QBrush(QColor(0, 0, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Inactive, QPalette.Text, brush)
        brush = QBrush(QColor(255, 255, 255))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Inactive, QPalette.BrightText, brush)
        brush = QBrush(QColor(0, 0, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Inactive, QPalette.ButtonText, brush)
        brush = QBrush(QColor(255, 255, 255))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Inactive, QPalette.Base, brush)
        brush = QBrush(QColor(255, 170, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Inactive, QPalette.Window, brush)
        brush = QBrush(QColor(0, 0, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Inactive, QPalette.Shadow, brush)
        brush = QBrush(QColor(255, 212, 127))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Inactive, QPalette.AlternateBase, brush)
        brush = QBrush(QColor(255, 255, 220))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Inactive, QPalette.ToolTipBase, brush)
        brush = QBrush(QColor(0, 0, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Inactive, QPalette.ToolTipText, brush)
        brush = QBrush(QColor(0, 0, 0, 128))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Inactive, QPalette.PlaceholderText, brush)
        brush = QBrush(QColor(127, 85, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Disabled, QPalette.WindowText, brush)
        brush = QBrush(QColor(255, 170, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Disabled, QPalette.Button, brush)
        brush = QBrush(QColor(255, 213, 127))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Disabled, QPalette.Light, brush)
        brush = QBrush(QColor(255, 191, 63))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Disabled, QPalette.Midlight, brush)
        brush = QBrush(QColor(127, 85, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Disabled, QPalette.Dark, brush)
        brush = QBrush(QColor(170, 113, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Disabled, QPalette.Mid, brush)
        brush = QBrush(QColor(127, 85, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Disabled, QPalette.Text, brush)
        brush = QBrush(QColor(255, 255, 255))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Disabled, QPalette.BrightText, brush)
        brush = QBrush(QColor(127, 85, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Disabled, QPalette.ButtonText, brush)
        brush = QBrush(QColor(255, 170, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Disabled, QPalette.Base, brush)
        brush = QBrush(QColor(255, 170, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Disabled, QPalette.Window, brush)
        brush = QBrush(QColor(0, 0, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Disabled, QPalette.Shadow, brush)
        brush = QBrush(QColor(255, 170, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Disabled, QPalette.AlternateBase, brush)
        brush = QBrush(QColor(255, 255, 220))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Disabled, QPalette.ToolTipBase, brush)
        brush = QBrush(QColor(0, 0, 0))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Disabled, QPalette.ToolTipText, brush)
        brush = QBrush(QColor(0, 0, 0, 128))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Disabled, QPalette.PlaceholderText, brush)
        self.setPalette(palette)
